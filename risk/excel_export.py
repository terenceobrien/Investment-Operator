"""
Excel export for the portfolio risk model.

Writes a single multi-tab workbook summarizing everything the beta and factor
models produce. Designed to be called by both run_beta.py and run_factors.py,
and structured so the data stays manipulable (clean tables, real numbers,
consistent formatting) for later analysis.

Tabs:
  Summary            - headline betas, factor exposures, stress drawdown
  Positions          - ingested weights, values, cash breakdown
  Factor Loadings    - per-name factor betas + r2 + specific var
  Risk Decomposition - factor vs specific, per-factor risk contribution (raw+EWMA)
  Breadth            - effective independent bets and implied IR
  Stress Test        - stressed vols, implied drawdown, factor contributions
  Factor Covariance  - the F matrix + correlation matrix

Formatting follows the model conventions: Inter-style clean headers, mono-ish
numeric alignment, purple accent to match the Helix palette, negatives in red.
"""

from __future__ import annotations

from datetime import datetime
from math import ceil, log10
from numbers import Real

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FACTOR_ORDER = ["MKT", "AI", "MOM", "QUAL", "VAL", "SIZE", "LOWVOL"]
BETA_FORMAT = "__BETA_FORMAT__"

# Helix-ish palette
ACCENT = "9580D4"       # purple
HEADER_BG = "1A1A22"    # near-black header
HEADER_FG = "FFFFFF"
GREEN = "57A06A"
RED = "B85555"
HAIRLINE = "3A3A44"
SUBTLE = "6A6A78"

FONT = "Arial"          # skill-preferred professional font

_thin = Side(style="thin", color=HAIRLINE)
_border = Border(bottom=_thin)


def _style_header(cell):
    cell.font = Font(name=FONT, bold=True, color=HEADER_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border


def _is_number(value) -> bool:
    return isinstance(value, Real) and not isinstance(value, bool) and not pd.isna(value)


def _clean_cell_value(value):
    if pd.isna(value):
        return None
    return value


def _beta_number_format(value) -> str:
    if value is None or not _is_number(value):
        return "0.00"
    mag = abs(float(value))
    if mag < 1e-6:
        return "0.00"
    if mag < 0.005:
        decimals = max(2, int(ceil(-log10(mag))))
        return "0." + ("0" * decimals)
    return "0.00"


def _apply_number_format(cell, value, fmt):
    if not fmt:
        return
    if fmt == BETA_FORMAT:
        cell.number_format = _beta_number_format(value)
    else:
        cell.number_format = fmt


def _title(ws, row, text, span=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=13, color=ACCENT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _subtle(ws, row, text, span=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, italic=True, size=9, color=SUBTLE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _write_df(
    ws,
    df,
    start_row,
    start_col=1,
    pct_cols=None,
    num_fmt="0.000",
    pct_fmt="0.0%",
    money_cols=None,
    beta_cols=None,
    format_map=None,
):
    """Write a DataFrame with header styling and per-column number formats."""
    pct_cols = pct_cols or set()
    money_cols = money_cols or set()
    beta_cols = beta_cols or set()
    format_map = format_map or {}
    # header
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=str(col))
        _style_header(cell)
    # body
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            val = _clean_cell_value(row[col])
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            cell.font = Font(name=FONT, size=10)
            if _is_number(val):
                if col in pct_cols:
                    _apply_number_format(cell, val, pct_fmt)
                elif col in money_cols:
                    _apply_number_format(cell, val, '$#,##0.00')
                elif col in beta_cols:
                    _apply_number_format(cell, val, BETA_FORMAT)
                elif col in format_map:
                    _apply_number_format(cell, val, format_map[col])
                else:
                    _apply_number_format(cell, val, num_fmt)
                if val < 0:
                    cell.font = Font(name=FONT, size=10, color=RED)
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    # autosize-ish
    for j, col in enumerate(df.columns):
        width = max(
            len(str(col)) + 2,
            max((len(str(_clean_cell_value(row[col]) or "")) for _, row in df.iterrows()), default=8) + 2,
        )
        ws.column_dimensions[get_column_letter(start_col + j)].width = min(width, 30)
    return start_row + 1 + len(df)


def _kv(ws, row, label, value, fmt=None, color=None, bold=False):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(name=FONT, size=10, bold=bold)
    value = _clean_cell_value(value)
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = Font(name=FONT, size=10, bold=bold, color=color or "000000")
    _apply_number_format(vc, value, fmt)
    vc.alignment = Alignment(horizontal="right")
    return row + 1


def _positions_frame(ingest) -> pd.DataFrame:
    pos_rows = []
    for tkr, w in sorted(ingest.weights.items(), key=lambda kv: -kv[1]):
        pos_rows.append({
            "Ticker": tkr,
            "Weight": w * ingest.invested_fraction,
            "Value ($)": ingest.per_symbol_value.get(tkr, 0.0),
        })
    pos_rows.append({
        "Ticker": "CASH",
        "Weight": ingest.cash_fraction,
        "Value ($)": ingest.cash_value,
    })
    return pd.DataFrame(pos_rows)


def export_workbook(
    path: str,
    *,
    ingest,                     # IngestResult
    beta=None,                  # dict of beta headline numbers (optional)
    factor=None,                # FactorResult (optional)
    risk_by_method=None,        # {method: FactorRiskResult} (optional)
    stress=None,                # StressResult (optional)
    breadth=None,               # BreadthResult (optional)
    generated_at: datetime | None = None,
):
    """Write the full workbook. Sections are optional so run_beta and run_factors
    can each populate what they have."""
    wb = Workbook()
    gen = generated_at or datetime.now()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    r = _title(ws, 1, "Portfolio Risk Summary")
    r = _subtle(ws, r, f"Generated {gen:%Y-%m-%d %H:%M}  |  estimation window per model tab")
    r += 1

    r = _kv(ws, r, "Total account value", ingest.total_account_value, '$#,##0.00', bold=True)
    r = _kv(ws, r, "Invested (equities)", ingest.invested_value, '$#,##0.00')
    r = _kv(ws, r, "Cash / core", ingest.cash_value, '$#,##0.00')
    r = _kv(ws, r, "Invested fraction", ingest.invested_fraction, '0.0%')
    r = _kv(ws, r, "Cash fraction", ingest.cash_fraction, '0.0%')
    r += 1

    if beta:
        r = _title(ws, r, "Beta")
        r = _kv(ws, r, "Invested-sleeve beta (shrunk)", beta.get("sleeve_shrunk"), BETA_FORMAT, bold=True)
        r = _kv(ws, r, "Invested-sleeve beta (raw)", beta.get("sleeve_raw"), BETA_FORMAT)
        r = _kv(ws, r, "Whole-book beta (shrunk)", beta.get("book_shrunk"), BETA_FORMAT, bold=True)
        r = _kv(ws, r, "Whole-book beta (raw)", beta.get("book_raw"), BETA_FORMAT)
        r += 1

    if factor:
        r = _title(ws, r, "Factor Exposures (invested sleeve)")
        for fac in [f for f in FACTOR_ORDER if f in factor.portfolio_exposures]:
            expo = factor.portfolio_exposures[fac]
            col = RED if expo < 0 else "000000"
            r = _kv(ws, r, fac, expo, BETA_FORMAT, color=col)
        r += 1

    if breadth:
        r = _title(ws, r, "Breadth")
        r = _kv(
            ws,
            r,
            "Effective independent bets",
            breadth.effective_n_eigen,
            '0.0',
            bold=True,
        )
        r = _kv(
            ws,
            r,
            "Effective annual breadth",
            breadth.effective_breadth_annual,
            '0.0',
        )
        r += 1

    if stress:
        r = _title(ws, r, "Stress Test")
        r = _subtle(ws, r, stress.shock_description)
        r = _kv(ws, r, "Stressed total vol (annual)", stress.stressed_total_vol_annual, '0.0%')
        r = _kv(ws, r, "Implied sleeve drawdown", stress.implied_drawdown, '0.0%',
                color=RED, bold=True)
        r = _kv(ws, r, "Implied whole-book drawdown",
                ingest.invested_fraction * stress.implied_drawdown, '0.0%',
                color=RED, bold=True)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    # ---------------- Positions ----------------
    ws = wb.create_sheet("Positions")
    r = _title(ws, 1, "Positions")
    r += 1
    _write_df(ws, _positions_frame(ingest), r, pct_cols={"Weight"}, money_cols={"Value ($)"})

    # ---------------- Factor Loadings ----------------
    if factor is not None:
        ws = wb.create_sheet("Factor Loadings")
        r = _title(ws, 1, "Per-Name Factor Loadings")
        r = _subtle(ws, r, f"Window: {factor.n_obs} periods  |  "
                           f"R2 mean {factor.r2_summary['mean']:.2f} / "
                           f"min {factor.r2_summary['min']:.2f}")
        r += 1
        ld = factor.per_ticker_loadings.copy()
        # order columns: ticker, factors, r2, specific_var
        cols = ["ticker"] + FACTOR_ORDER + ["r2", "specific_var"]
        ld = ld[[c for c in cols if c in ld.columns]]
        _write_df(
            ws,
            ld,
            r,
            pct_cols={"r2"},
            beta_cols=set(FACTOR_ORDER),
            format_map={"specific_var": "0.000000"},
        )

    # ---------------- Risk Decomposition ----------------
    if risk_by_method:
        ws = wb.create_sheet("Risk Decomposition")
        r = _title(ws, 1, "Factor Covariance Risk Decomposition")
        r += 1
        for method, rr in risk_by_method.items():
            tag = "RAW (sample cov)" if method == "sample" else "EWMA (recent-weighted)"
            r = _title(ws, r, tag, span=5)
            r = _kv(ws, r, "Total vol (annual)", rr.total_vol_annual, '0.0%', bold=True)
            r = _kv(ws, r, "Factor vol", rr.factor_vol_annual, '0.0%')
            r = _kv(ws, r, "Specific vol", rr.specific_vol_annual, '0.0%')
            r = _kv(ws, r, "Factor share of variance", rr.pct_factor, '0.0%')
            r = _kv(ws, r, "Specific share of variance", rr.pct_specific, '0.0%')
            r += 1
            rc = rr.risk_contributions.copy()
            rc = rc[["factor", "exposure", "risk_contribution",
                     "pct_of_total_var"]]
            r = _write_df(
                ws,
                rc,
                r,
                pct_cols={"pct_of_total_var"},
                beta_cols={"exposure"},
                format_map={"risk_contribution": "0.000000"},
            )
            r += 2

    # ---------------- Breadth ----------------
    if breadth is not None:
        ws = wb.create_sheet("Breadth")
        _write_breadth_sheet(ws, breadth)

    # ---------------- Stress Test ----------------
    if stress is not None:
        ws = wb.create_sheet("Stress Test")
        r = _title(ws, 1, "Stress Test")
        r = _subtle(ws, r, stress.shock_description)
        r += 1
        r = _kv(ws, r, "Base factor vol (annual)", stress.base_factor_vol_annual, '0.0%')
        r = _kv(ws, r, "Stressed factor vol", stress.stressed_factor_vol_annual, '0.0%')
        r = _kv(ws, r, "Stressed total vol", stress.stressed_total_vol_annual, '0.0%')
        r = _kv(ws, r, "Vol scale", stress.vol_scale, '0.0"x"')
        r = _kv(ws, r, "Correlation floor", stress.corr_floor, '0.00')
        r = _kv(ws, r, "Implied sleeve drawdown", stress.implied_drawdown, '0.0%',
                color=RED, bold=True)
        r += 1
        r = _title(ws, r, "Drawdown Contribution by Factor", span=4)
        _write_df(
            ws,
            stress.factor_shock_contributions,
            r,
            pct_cols={"cond_factor_move", "contribution_to_drawdown"},
            beta_cols={"raw_exposure"},
        )

    # ---------------- Factor Covariance ----------------
    if risk_by_method:
        ws = wb.create_sheet("Factor Covariance")
        # use the sample covariance/correlation
        rr = risk_by_method.get("sample") or next(iter(risk_by_method.values()))
        r = _title(ws, 1, "Factor Covariance Matrix (periodic)")
        r += 1
        cov = rr.cov.copy().reset_index().rename(columns={"index": ""})
        r = _write_df(ws, cov, r, num_fmt="0.000000")
        r += 2
        r = _title(ws, r, "Factor Correlation Matrix")
        r += 1
        corr = rr.corr.copy().reset_index().rename(columns={"index": ""})
        _write_df(ws, corr, r, num_fmt="0.00")

    _finalize_workbook(wb)
    wb.save(path)
    return path


def export_beta_workbook(
    path: str,
    *,
    ingest,
    beta,
    beta_summary,
    per_ticker=None,
    generated_at: datetime | None = None,
):
    """Write the styled beta workbook used by run_beta.py."""
    wb = Workbook()
    gen = generated_at or datetime.now()

    ws = wb.active
    ws.title = "Summary"
    r = _title(ws, 1, "Portfolio Beta Summary")
    r = _subtle(ws, r, f"Generated {gen:%Y-%m-%d %H:%M}")
    r += 1
    r = _kv(ws, r, "Total account value", ingest.total_account_value, '$#,##0.00', bold=True)
    r = _kv(ws, r, "Invested (equities)", ingest.invested_value, '$#,##0.00')
    r = _kv(ws, r, "Cash / core", ingest.cash_value, '$#,##0.00')
    r = _kv(ws, r, "Invested fraction", ingest.invested_fraction, '0.0%')
    r = _kv(ws, r, "Cash fraction", ingest.cash_fraction, '0.0%')
    r += 1

    r = _title(ws, r, "Beta")
    r = _subtle(ws, r, f"Market: {beta_summary.get('market')} | lookback {beta_summary.get('lookback')} periods")
    r = _kv(ws, r, "Invested-sleeve beta (shrunk)", beta_summary.get("sleeve_beta_shrunk"), BETA_FORMAT, bold=True)
    r = _kv(ws, r, "Invested-sleeve beta (raw)", beta_summary.get("sleeve_beta_raw"), BETA_FORMAT)
    r = _kv(ws, r, "Whole-book beta (shrunk)", beta_summary.get("whole_book_beta_shrunk"), BETA_FORMAT, bold=True)
    r = _kv(ws, r, "Whole-book beta (raw)", beta_summary.get("whole_book_beta_raw"), BETA_FORMAT)
    r = _kv(ws, r, "SPY position weight", beta_summary.get("spy_position_weight"), '0.0%')
    r = _kv(ws, r, "Non-SPY sleeve share", beta_summary.get("non_spy_share"), '0.0%')

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    ws = wb.create_sheet("Positions")
    r = _title(ws, 1, "Positions")
    r += 1
    _write_df(ws, _positions_frame(ingest), r, pct_cols={"Weight"}, money_cols={"Value ($)"})

    ws = wb.create_sheet("Per Name Beta")
    r = _title(ws, 1, "Per-Name Beta")
    r = _subtle(
        ws,
        r,
        f"Observations: {getattr(beta, 'n_obs', 'n/a')} | halflife {getattr(beta, 'halflife', 'n/a')} | prior {getattr(beta, 'prior_mean', 'n/a')}",
        span=8,
    )
    r += 1
    df = per_ticker if per_ticker is not None else beta.per_ticker
    beta_cols = {"raw_beta", "shrunk_beta", "beta_delta", "raw_beta_var", "contrib_raw", "contrib_shrunk", "contribution"}
    _write_df(
        ws,
        df.copy(),
        r,
        pct_cols={"weight"},
        beta_cols=beta_cols,
        format_map={"raw_beta_var": "0.000000"},
    )

    _finalize_workbook(wb)
    wb.save(path)
    return path


def _write_breadth_sheet(ws, breadth) -> None:
    r = _title(ws, 1, "Portfolio Breadth (independent bets)", span=4)
    r = _subtle(
        ws,
        r,
        "Effective N (eigenvalue) is the reliable measure; avg-corr is a conservative floor.",
        span=4,
    )
    r += 1
    for label, value, fmt, bold in (
        ("Positions", breadth.n_positions, '0', False),
        ("Avg pairwise correlation", breadth.avg_pairwise_corr, '0.00', False),
        ("Effective N (avg-corr)", breadth.effective_n_avgcorr, '0.0', False),
        ("Effective N (eigenvalue)", breadth.effective_n_eigen, '0.0', True),
        ("Concentration ratio", breadth.concentration_ratio, '0.0%', False),
        ("Top principal component", breadth.top_eigen_share, '0.0%', False),
        ("Decisions per year", breadth.decisions_per_year, '0.0', False),
        ("Effective annual breadth", breadth.effective_breadth_annual, '0.0', False),
    ):
        r = _kv(ws, r, label, value, fmt, bold=bold)

    r += 2
    r = _title(
        ws,
        r,
        "Implied Information Ratio (Fundamental Law IR = IC x sqrt(BR))",
        span=4,
    )
    r = _subtle(
        ws,
        r,
        "Reference: good IC 0.05, great 0.10, world-class 0.15; IR>1.0 reached by ~10% of managers.",
        span=4,
    )
    r += 1
    ir_rows = [
        {"IC (skill)": float(ic), "Implied IR": float(ir)}
        for ic, ir in sorted(breadth.implied_ir_at_ic.items(), key=lambda item: float(item[0]))
    ]
    _write_df(
        ws,
        pd.DataFrame(ir_rows),
        r,
        format_map={"IC (skill)": "0.00", "Implied IR": "0.00"},
    )

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18


def _finalize_workbook(wb) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass
    _confirm_zero_formula_errors(wb)


def _confirm_zero_formula_errors(wb) -> None:
    formula_cells = []
    error_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "e":
                    error_cells.append(f"{ws.title}!{cell.coordinate}")
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(f"{ws.title}!{cell.coordinate}")
    if error_cells:
        raise ValueError(f"Workbook contains formula error cells: {error_cells}")
    if formula_cells:
        raise ValueError(f"Workbook contains formulas despite value-only export: {formula_cells}")
