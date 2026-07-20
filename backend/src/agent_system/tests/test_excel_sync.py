from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook

from src.agent_system.schemas.trade_outcome import TradeOutcome
from src.agent_system.services import excel_sync
from src.agent_system.services.excel_sync import ExcelSync
from src.agent_system.storage.repository import DecisionLogEntry


CYCLE_ID = "8b1dcfdb-1dd2-4ccf-8afa-be2ab2dae01e"


def _short_cycle_id(cycle_id: str) -> str:
    return cycle_id[-8:]


def _outcome() -> TradeOutcome:
    return TradeOutcome(
        trade_id="trade-accepted",
        cycle_id=CYCLE_ID,
        cycle_date="2026-06-25",
        underlying="ACPT",
        priority_theme="Accepted test theme",
        direction="long",
        instrument_type="stock",
        instrument_description="ACPT common stock",
        proposed_size_pct=0.02,
        final_size_pct=0.02,
        decision="execute",
        variant_strength="moderate",
        conviction="moderate",
    )


def _shadow_outcome() -> TradeOutcome:
    return TradeOutcome(
        trade_id="trade-rejected",
        cycle_id=CYCLE_ID,
        cycle_date="2026-06-25",
        underlying="RJCT",
        direction="long",
        instrument_type="single_stock",
        instrument_description="shadow tracking of RJCT",
        proposed_size_pct=0.0,
        final_size_pct=0.0,
        decision="shadow_rejected",
        conviction="weak",
        status="shadow_rejected",
        entry_triggered=False,
        entry_date="2026-06-25",
        entry_underlying_price=100.0,
        current_underlying_price=110.0,
        current_unrealized_pnl_pct=0.10,
        days_held=5,
        audit_notes="SYSTEM REJECTED: weak_rule: not enough evidence",
    )


def _rejected_entry() -> DecisionLogEntry:
    return DecisionLogEntry(
        id=1,
        cycle_id=CYCLE_ID,
        candidate="RJCT",
        decision="rejected",
        conviction_rating="weak",
        rule_applied="weak_rule",
        summary="Rejected because the variant evidence was not strong enough.",
        weakest_link="fundamental",
        trade_idea_id="trade-rejected",
        timestamp=datetime(2026, 6, 25, tzinfo=timezone.utc),
        created_at=datetime(2026, 6, 25, tzinfo=timezone.utc),
    )


def _workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Detailed Log"
    sheet["A1"] = "Cycle date"
    sheet["B1"] = "Cycle ID"
    sheet["D1"] = "Ticker"
    sheet["A3"] = "Date the cycle that produced the trade ran"
    sheet["B3"] = "Last 8 chars of cycle UUID is fine"
    sheet["D3"] = "Underlying"
    sheet["A4"] = datetime(2026, 6, 25)
    sheet["B4"] = _short_cycle_id(CYCLE_ID)
    sheet["D4"] = "ACPT"
    main = workbook.create_sheet("Main Log")
    for column, header in {
        "A": "Cycle date",
        "B": "Cycle ID",
        "C": "Priority theme",
        "D": "Ticker",
        "E": "Instrument",
        "F": "Entry date",
        "G": "Entry price",
        "H": "Held days",
        "I": "Current price",
        "J": "Current underlying",
        "K": "Last price update",
        "L": "Unrealized P&L %",
        "M": "Pipeline verdict",
    }.items():
        main[f"{column}1"] = header
    main["A2"] = datetime(2026, 6, 25)
    main["B2"] = _short_cycle_id(CYCLE_ID)
    main["D2"] = "ACPT"
    workbook.save(path)


def test_excel_sync_appends_rejected_candidates_to_main_log(
    tmp_path,
    monkeypatch,
):
    path = tmp_path / "trade_log.xlsx"
    _workbook(path)
    outcome = _outcome()
    rejected = _rejected_entry()

    monkeypatch.setattr(excel_sync, "load_trade_outcomes", lambda: [outcome])
    monkeypatch.setattr(
        excel_sync,
        "load_decision_log_entries_by_cycle",
        lambda cycle_id: [rejected] if cycle_id == CYCLE_ID else [],
    )
    monkeypatch.setattr(excel_sync, "save_trade_outcome", lambda _outcome: None)

    report = ExcelSync(path, log_warnings=False).sync()

    assert len(report.rejected_candidates_appended) == 1
    assert report.rejected_candidates_appended[0].underlying == "RJCT"

    workbook = load_workbook(path)
    sheet = workbook["Detailed Log"]
    main = workbook["Main Log"]
    assert sheet["AE1"].value is None
    assert sheet["AF1"].value is None
    assert sheet["B5"].value == _short_cycle_id(CYCLE_ID)
    assert sheet["D5"].value == "RJCT"
    assert sheet["H5"].value == "weak"
    assert sheet["AE5"].value is None
    assert sheet["AF5"].value is None
    assert main["B3"].value == _short_cycle_id(CYCLE_ID)
    assert main["D3"].value == "RJCT"
    assert main["M3"].value == "REJECTED"

    second_report = ExcelSync(path, log_warnings=False).sync()

    assert second_report.rejected_candidates_appended == []


def test_excel_sync_updates_existing_shadow_rejected_row(tmp_path, monkeypatch):
    path = tmp_path / "trade_log.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    sheet = workbook["Detailed Log"]
    sheet["B5"] = _short_cycle_id(CYCLE_ID)
    sheet["D5"] = "RJCT"
    sheet["Z5"] = "legacy current price"
    sheet["AA5"] = "legacy current underlying"
    sheet["AB5"] = "legacy last update"
    sheet["AC5"] = "legacy pnl"
    sheet["AE5"] = "legacy verdict"
    sheet["AF5"] = "legacy reason"
    main = workbook["Main Log"]
    main["B3"] = _short_cycle_id(CYCLE_ID)
    main["D3"] = "RJCT"
    main["L3"] = '=IFERROR((I3-G3)/G3,"")'
    workbook.save(path)
    outcome = _outcome()
    shadow = _shadow_outcome()

    monkeypatch.setattr(excel_sync, "load_trade_outcomes", lambda: [outcome, shadow])
    monkeypatch.setattr(
        excel_sync,
        "load_decision_log_entries_by_cycle",
        lambda _cycle_id: [],
    )
    monkeypatch.setattr(excel_sync, "save_trade_outcome", lambda _outcome: None)

    report = ExcelSync(path, log_warnings=False).sync()

    assert report.rejected_candidates_appended == []
    assert not any(
        write.field_name == "current_unrealized_pnl_pct"
        for write in report.system_fields_written
    )
    workbook = load_workbook(path)
    sheet = workbook["Detailed Log"]
    main = workbook["Main Log"]
    assert sheet["P5"].value.date().isoformat() == "2026-06-25"
    assert sheet["Q5"].value == 100.0
    assert sheet["Z5"].value == "legacy current price"
    assert sheet["AA5"].value == "legacy current underlying"
    assert sheet["AB5"].value == "legacy last update"
    assert sheet["AC5"].value == "legacy pnl"
    assert sheet["O5"].value == "NO (shadow)"
    assert sheet["AE5"].value == "legacy verdict"
    assert sheet["AF5"].value == "legacy reason"
    assert main["G3"].value == 100.0
    assert main["H3"].value == 5
    assert main["I3"].value == 110.0
    assert main["J3"].value == 110.0
    assert main["L3"].value == '=IFERROR((I3-G3)/G3,"")'
    assert main["M3"].value == "REJECTED"
