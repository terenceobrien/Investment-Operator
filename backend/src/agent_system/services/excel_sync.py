"""Bidirectional sync between TradeOutcome JSONL storage and helix_trade_log.xlsx.

The sync model is column-based:
- SYSTEM_COLUMNS are written by sync from TradeOutcome -> Excel. Manual edits
  will be overwritten.
- USER_COLUMNS are written by sync from Excel -> TradeOutcome. User edits are
  preserved and reflected in storage.

Run order on each sync:
1. Read Excel - capture user-owned columns into TradeOutcome.
2. Write Excel - write system-owned columns from TradeOutcome.

This means the user's edits are always read first before any system overwrites
happen.
"""
from __future__ import annotations

import logging
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Optional

from src.agent_system.schemas.trade_outcome import TradeOutcome
from src.agent_system.storage.repository import (
    DecisionLogEntry,
    load_decision_log_entries_by_cycle,
    load_trade_outcomes,
    save_trade_outcome,
)
from src.agent_system.paths import agent_system_data_root

logger = logging.getLogger(__name__)


EXCEL_LOG_PATH = agent_system_data_root(create=False) / "helix_trade_log.xlsx"
DATA_START_ROW = 3
MAIN_LOG_DATA_START_ROW = 2
HEADER_ROW = 1
DESCRIPTION_ROW = 2
TRADE_LOG_SHEET = "Detailed Log"
MAIN_LOG_SHEET = "Main Log"


NEW_COLUMNS_TO_INITIALIZE = {
    # Restored after the original "Z" audit-notes column was deleted, which had
    # shifted every column above one slot to the left of where this code
    # expected them. Given a new home at AG rather than reclaiming Z, since Z
    # is now the live "Current price" column.
    "AG": {
        "header": "Audit notes",
        "description": "Free-text notes; user-editable, never overwritten by sync",
    },
}


MAIN_LOG_HEADERS = {
    "A": "Cycle date",
    "B": "Cycle ID",
    "C": "Priority theme",
    "D": "Ticker",
    "E": "Instrument",
    "F": "Entry date",
    "G": "Entry price",
    "H": "Held days",
    "I": "Current price",
    "J": "Current underlying",
    "K": "Last price update",
    "M": "Pipeline verdict",
}


MAIN_LOG_COLUMN_FORMATS = {
    "I": "$#,##0.00",
    "J": "$#,##0.00",
    "K": "yyyy-mm-dd",
}


# System-owned: written by sync, sourced from TradeOutcome.
SYSTEM_COLUMNS = {
    "A": "cycle_date",
    "B": "cycle_id",
    "C": "priority_theme",
    "D": "underlying",
    "E": "direction",
    "F": "instrument_description",
    "G": "variant_strength",
    "H": "conviction",
    "I": "robustness_score",
    "J": "robustness_quartile",
    "K": "proposed_size_pct",
    "L": "final_size_pct",
    "P": "entry_date",
    "Q": "entry_instrument_price",
    "R": "exit_date",
    "S": "exit_instrument_price",
    # NOTE: As of this update, column U (P&L %) holds realized P&L only.
    # Unrealized/current price and pipeline verdict fields now live only on
    # Main Log. Existing Detailed Log values in old columns are left untouched.
    "U": "realized_pnl_pct",
}


MAIN_LOG_COLUMNS = {
    "A": "cycle_date",
    "B": "cycle_id",
    "C": "priority_theme",
    "D": "underlying",
    "E": "instrument_description",
    "F": "entry_date",
    "G": "entry_instrument_price",
    "H": "days_held",
    "I": "current_price_display",
    "J": "current_underlying_price",
    "K": "last_price_update",
    "M": "pipeline_verdict",
}


# User-owned: read by sync, written to TradeOutcome.
USER_COLUMNS = {
    "M": "user_decision",
    "N": "user_decision_reason",
    "O": "entry_triggered",
    "T": "exit_reason",
    "W": "thesis_played_out",
    "X": "win_source",
    "Y": "system_contribution",
    "AG": "audit_notes",
}


DATE_COLUMNS = {"A", "P", "R"}
PERCENT_COLUMNS = {"K", "L", "U"}
MAIN_LOG_DATE_COLUMNS = {"A", "F", "K"}
MAIN_LOG_PERCENT_COLUMNS: set[str] = set()
ALL_LOG_COLUMNS = range(1, 34)  # A..AG
MAIN_LOG_COLUMNS_RANGE = range(1, 14)  # A..M
MAIN_LOG_OWNED_COLUMN_NUMBERS = tuple(
    column for column in MAIN_LOG_COLUMNS_RANGE if column != 12
)


@dataclass
class UserEditApplied:
    underlying: str
    cycle_id_short: str
    field_name: str
    old_value: Any
    new_value: Any
    status_before: str
    status_after: str


@dataclass
class SystemFieldWritten:
    underlying: str
    cycle_id_short: str
    column: str
    field_name: str
    old_value: Any
    new_value: Any
    row_index: int


@dataclass
class OutcomeAppended:
    underlying: str
    cycle_id_short: str
    row_index: int


@dataclass
class SyncWarning:
    underlying: str | None
    cycle_id_short: str | None
    message: str


@dataclass
class SyncConflict:
    underlying: str
    cycle_id_short: str
    field_name: str
    storage_value: Any
    excel_value: Any


@dataclass
class SystemOverwrite:
    underlying: str
    cycle_id_short: str
    field_name: str
    excel_value: Any
    storage_value: Any


@dataclass
class SyncReport:
    log_path: Path
    dry_run: bool = False
    user_edits_applied: list[UserEditApplied] = field(default_factory=list)
    system_fields_written: list[SystemFieldWritten] = field(default_factory=list)
    outcomes_appended: list[OutcomeAppended] = field(default_factory=list)
    rejected_candidates_appended: list[OutcomeAppended] = field(default_factory=list)
    warnings: list[SyncWarning] = field(default_factory=list)
    conflicts: list[SyncConflict] = field(default_factory=list)
    system_overwrites: list[SystemOverwrite] = field(default_factory=list)
    ignored_user_edits: int = 0


def _short_cycle_id(cycle_id: str | None) -> str:
    return str(cycle_id or "")[-8:]


def _key(cycle_id_short: str | None, underlying: str | None) -> tuple[str, str]:
    return (_cell_text(cycle_id_short).lower(), _cell_text(underlying).upper())


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _date_value(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _datetime_date_value(value: datetime | str | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    return _date_value(str(value))


def _resolve_current_price_display(outcome: TradeOutcome) -> Optional[float]:
    if outcome.current_instrument_price is not None:
        return outcome.current_instrument_price
    return outcome.current_underlying_price


class ExcelSync:
    def __init__(self, log_path: Path = EXCEL_LOG_PATH, *, log_warnings: bool = True):
        self.log_path = log_path
        self.log_warnings = log_warnings
        if not self.log_path.exists():
            raise FileNotFoundError(f"Excel log not found at {self.log_path}")

        self.workbook = None
        self.sheet = None
        self.main_sheet = None
        self._dry_run = False
        self._report: SyncReport | None = None
        self._outcomes: list[TradeOutcome] = []
        self._outcome_by_key: dict[tuple[str, str], TradeOutcome] = {}
        self._row_by_key: dict[tuple[str, str], int] = {}
        self._all_row_by_key: dict[tuple[str, str], int] = {}
        self._main_row_by_key: dict[tuple[str, str], int] = {}
        self._main_all_row_by_key: dict[tuple[str, str], int] = {}
        self._ambiguous_outcome_keys: set[tuple[str, str]] = set()
        self._warning_keys: set[tuple[str | None, str | None, str]] = set()

    def sync(self, dry_run: bool = False) -> SyncReport:
        """Run a full bidirectional sync."""
        from openpyxl import load_workbook

        self.workbook = load_workbook(self.log_path)
        if TRADE_LOG_SHEET not in self.workbook.sheetnames:
            raise KeyError(f"Workbook has no {TRADE_LOG_SHEET!r} sheet")
        if MAIN_LOG_SHEET not in self.workbook.sheetnames:
            raise KeyError(f"Workbook has no {MAIN_LOG_SHEET!r} sheet")
        self.sheet = self.workbook[TRADE_LOG_SHEET]
        self.main_sheet = self.workbook[MAIN_LOG_SHEET]
        self._dry_run = dry_run
        self._report = SyncReport(log_path=self.log_path, dry_run=dry_run)
        self._warning_keys = set()

        self.ensure_extended_columns(self.sheet)
        self.ensure_main_log_columns(self.main_sheet)
        self._refresh_indexes()
        self._report.user_edits_applied = self.read_user_edits()
        self._refresh_indexes()
        self._report.system_fields_written = [
            *self.write_system_fields(),
        ]
        self._report.outcomes_appended = self.append_new_outcomes()
        self._report.rejected_candidates_appended = self.append_rejected_candidates()
        self._refresh_indexes()
        self._ensure_main_log_rows_for_outcomes()
        self._refresh_indexes()
        self._report.system_fields_written.extend(self.write_main_log_fields())

        if not dry_run:
            self.workbook.save(self.log_path)

        report = self._report
        self._report = None
        return report

    def ensure_extended_columns(self, worksheet) -> None:
        """Write new headers/descriptions if absent; safe to call repeatedly."""
        for col_letter, fields in NEW_COLUMNS_TO_INITIALIZE.items():
            header_cell = worksheet[f"{col_letter}{HEADER_ROW}"]
            desc_cell = worksheet[f"{col_letter}{DESCRIPTION_ROW}"]
            if header_cell.value is None:
                header_cell.value = fields["header"]
                self._copy_cell_display_style(worksheet["A1"], header_cell)
            if desc_cell.value is None:
                desc_cell.value = fields["description"]
                self._copy_cell_display_style(worksheet["A2"], desc_cell)

    def ensure_main_log_columns(self, worksheet) -> None:
        """Write Main Log headers if absent; safe to call repeatedly."""

        for col_letter, expected_header in MAIN_LOG_HEADERS.items():
            header_cell = worksheet[f"{col_letter}{HEADER_ROW}"]
            if header_cell.value is not None:
                continue
            header_cell.value = expected_header
            self._copy_cell_display_style(worksheet["A1"], header_cell)

    def read_user_edits(self) -> list[UserEditApplied]:
        """Pass 1: pull USER_COLUMNS from Excel and apply to TradeOutcome records."""

        edits: list[UserEditApplied] = []
        now = datetime.now(timezone.utc)
        for key, row_index in sorted(self._row_by_key.items(), key=lambda item: item[1]):
            outcome = self._outcome_by_key.get(key)
            if outcome is None:
                continue
            if outcome.status == "shadow_rejected":
                continue

            status_before = outcome.status
            updates: dict[str, Any] = {}
            row_edits: list[UserEditApplied] = []

            for column, field_name in USER_COLUMNS.items():
                cell_value = self.sheet[f"{column}{row_index}"].value
                # V1 treats blank user-owned cells as "no change"; explicit clearing
                # should go through the CLI until we have deletion audit semantics.
                if _is_empty(cell_value):
                    continue

                normalized = self._normalize_user_value(
                    field_name,
                    cell_value,
                    outcome=outcome,
                    row_index=row_index,
                )
                if normalized is _INVALID:
                    self._report.ignored_user_edits += 1
                    continue

                old_value = getattr(outcome, field_name)
                if self._values_equal(field_name, old_value, normalized):
                    continue

                if not _is_empty(old_value):
                    self._report.conflicts.append(
                        SyncConflict(
                            underlying=outcome.underlying,
                            cycle_id_short=_short_cycle_id(outcome.cycle_id),
                            field_name=field_name,
                            storage_value=old_value,
                            excel_value=normalized,
                        )
                    )

                updates[field_name] = normalized
                row_edits.append(
                    UserEditApplied(
                        underlying=outcome.underlying,
                        cycle_id_short=_short_cycle_id(outcome.cycle_id),
                        field_name=field_name,
                        old_value=old_value,
                        new_value=normalized,
                        status_before=status_before,
                        status_after=status_before,
                    )
                )

            if not updates:
                continue

            self._apply_status_transitions(
                outcome=outcome,
                updates=updates,
                row_index=row_index,
                now=now,
            )
            updates["updated_at"] = now
            updated = outcome.model_copy_validate(updates)
            status_after = updated.status
            for edit in row_edits:
                edit.status_after = status_after
            edits.extend(row_edits)

            if not self._dry_run:
                save_trade_outcome(updated)

        return edits

    def write_system_fields(self) -> list[SystemFieldWritten]:
        """Pass 2: write SYSTEM_COLUMNS from TradeOutcome records to Excel."""

        writes: list[SystemFieldWritten] = []
        for key, row_index in sorted(self._row_by_key.items(), key=lambda item: item[1]):
            outcome = self._outcome_by_key.get(key)
            if outcome is None:
                continue
            writes.extend(self._write_system_fields_for_row(outcome, row_index))
        return writes

    def write_main_log_fields(self) -> list[SystemFieldWritten]:
        """Write Main Log system-owned summary, price, and pipeline fields."""

        writes: list[SystemFieldWritten] = []
        for key, row_index in sorted(
            self._main_row_by_key.items(),
            key=lambda item: item[1],
        ):
            outcome = self._outcome_by_key.get(key)
            if outcome is None:
                continue
            writes.extend(self._write_main_log_fields_for_row(outcome, row_index))
        return writes

    def append_new_outcomes(self) -> list[OutcomeAppended]:
        """Pass 3: append TradeOutcome records that have no Excel row yet."""

        appended: list[OutcomeAppended] = []
        row_index = self._first_empty_data_row()
        existing_keys = set(self._all_row_by_key)
        for outcome in sorted(
            [outcome for outcome in self._outcomes if outcome.status != "shadow_rejected"],
            key=lambda item: (
                item.cycle_date,
                _short_cycle_id(item.cycle_id),
                item.underlying,
            ),
        ):
            key = self._outcome_key(outcome)
            if key in existing_keys or key in self._ambiguous_outcome_keys:
                continue

            if not self._dry_run:
                self._copy_row_style(self._template_row(), row_index)
                for column in USER_COLUMNS:
                    self.sheet[f"{column}{row_index}"].value = None
                self._write_system_fields_for_row(outcome, row_index)
                self._append_main_log_outcome_row(outcome)

            appended.append(
                OutcomeAppended(
                    underlying=outcome.underlying,
                    cycle_id_short=_short_cycle_id(outcome.cycle_id),
                    row_index=row_index,
                )
            )
            existing_keys.add(key)
            self._all_row_by_key[key] = row_index
            if not self._dry_run:
                self._row_by_key[key] = row_index
            row_index += 1
        return appended

    def append_rejected_candidates(self) -> list[OutcomeAppended]:
        """Append system-rejected conviction-gate candidates to the trade log."""

        appended: list[OutcomeAppended] = []
        cycle_outcome: dict[str, TradeOutcome] = {}
        for outcome in self._outcomes:
            cycle_outcome.setdefault(outcome.cycle_id, outcome)

        row_index = self._first_empty_data_row()
        if self._dry_run and self._report is not None:
            row_index += len(self._report.outcomes_appended)

        existing_keys = set(self._all_row_by_key)
        if self._report is not None:
            for outcome_appended in self._report.outcomes_appended:
                existing_keys.add(
                    _key(
                        outcome_appended.cycle_id_short,
                        outcome_appended.underlying,
                    )
                )

        for outcome in sorted(
            [outcome for outcome in self._outcomes if outcome.status == "shadow_rejected"],
            key=lambda item: (
                item.cycle_date,
                _short_cycle_id(item.cycle_id),
                item.underlying,
            ),
        ):
            key = self._outcome_key(outcome)
            if key in existing_keys or key in self._ambiguous_outcome_keys:
                continue
            if not self._dry_run:
                self._copy_row_style(self._template_row(), row_index)
                for column in USER_COLUMNS:
                    self.sheet[f"{column}{row_index}"].value = None
                self._write_system_fields_for_row(outcome, row_index)
                self._append_main_log_outcome_row(outcome)

            appended.append(
                OutcomeAppended(
                    underlying=outcome.underlying,
                    cycle_id_short=_short_cycle_id(outcome.cycle_id),
                    row_index=row_index,
                )
            )
            existing_keys.add(key)
            self._all_row_by_key[key] = row_index
            if not self._dry_run:
                self._row_by_key[key] = row_index
            row_index += 1

        for cycle_id, representative in sorted(
            cycle_outcome.items(),
            key=lambda item: (item[1].cycle_date, _short_cycle_id(item[0])),
        ):
            # Screen-eliminated candidates are intentionally not represented here;
            # only candidates that reached the conviction gate have decision_log rows.
            rejected_entries = [
                entry
                for entry in load_decision_log_entries_by_cycle(cycle_id)
                if entry.decision == "rejected"
            ]
            for entry in sorted(rejected_entries, key=lambda item: item.candidate):
                key_variants = self._decision_log_key_variants(entry)
                if existing_keys.intersection(key_variants):
                    continue

                if not self._dry_run:
                    self._copy_row_style(self._template_row(), row_index)
                    self._write_rejected_candidate_row(
                        entry,
                        row_index,
                        representative,
                    )
                    self._append_main_log_rejected_candidate_row(
                        entry,
                        representative,
                    )

                appended.append(
                    OutcomeAppended(
                        underlying=entry.candidate,
                        cycle_id_short=_short_cycle_id(entry.cycle_id),
                        row_index=row_index,
                    )
                )
                existing_keys.update(key_variants)
                self._all_row_by_key[
                    _key(_short_cycle_id(entry.cycle_id), entry.candidate)
                ] = row_index
                row_index += 1
        return appended

    def _refresh_indexes(self) -> None:
        self._outcomes = load_trade_outcomes()
        self._outcome_by_key = {}
        self._ambiguous_outcome_keys = set()
        for outcome in self._outcomes:
            key = self._outcome_key(outcome)
            if key in self._outcome_by_key:
                self._ambiguous_outcome_keys.add(key)
                self._add_warning(
                    outcome.underlying,
                    _short_cycle_id(outcome.cycle_id),
                    "multiple TradeOutcome records match this Excel key; skipping sync for that key",
                )
                continue
            self._outcome_by_key[key] = outcome
        for key in self._ambiguous_outcome_keys:
            self._outcome_by_key.pop(key, None)

        self._row_by_key = {}
        self._all_row_by_key = {}
        for row_index in range(DATA_START_ROW, self.sheet.max_row + 1):
            if not self._is_trade_log_data_row(row_index):
                continue
            cycle_short = _cell_text(self.sheet[f"B{row_index}"].value)
            underlying = _cell_text(self.sheet[f"D{row_index}"].value)
            key = _key(cycle_short, underlying)
            if key in self._all_row_by_key:
                self._add_warning(
                    underlying or None,
                    cycle_short or None,
                    f"duplicate Excel row key at row {row_index}; leaving duplicate row untouched",
                )
                continue
            self._all_row_by_key[key] = row_index
            outcome = self._outcome_by_key.get(key)
            if key not in self._outcome_by_key:
                self._add_warning(
                    underlying or None,
                    cycle_short or None,
                    "found in Excel but no matching TradeOutcome - left untouched",
                )
                continue
            self._row_by_key[key] = row_index

        self._main_row_by_key = {}
        self._main_all_row_by_key = {}
        for row_index in range(MAIN_LOG_DATA_START_ROW, self.main_sheet.max_row + 1):
            if not self._is_main_log_data_row(row_index):
                continue
            cycle_short = _cell_text(self.main_sheet[f"B{row_index}"].value)
            underlying = _cell_text(self.main_sheet[f"D{row_index}"].value)
            key = _key(cycle_short, underlying)
            if key in self._main_all_row_by_key:
                self._add_warning(
                    underlying or None,
                    cycle_short or None,
                    f"duplicate Main Log row key at row {row_index}; leaving duplicate row untouched",
                )
                continue
            self._main_all_row_by_key[key] = row_index
            if key in self._outcome_by_key:
                self._main_row_by_key[key] = row_index

    def _write_system_fields_for_row(
        self,
        outcome: TradeOutcome,
        row_index: int,
    ) -> list[SystemFieldWritten]:
        writes: list[SystemFieldWritten] = []
        for column, field_name in SYSTEM_COLUMNS.items():
            cell = self.sheet[f"{column}{row_index}"]
            desired = self._system_value(outcome, column, field_name)
            if self._cell_matches_system_value(column, cell.value, desired):
                continue

            if not _is_empty(cell.value):
                self._report.system_overwrites.append(
                    SystemOverwrite(
                        underlying=outcome.underlying,
                        cycle_id_short=_short_cycle_id(outcome.cycle_id),
                        field_name=field_name,
                        excel_value=cell.value,
                        storage_value=desired,
                    )
                )

            writes.append(
                SystemFieldWritten(
                    underlying=outcome.underlying,
                    cycle_id_short=_short_cycle_id(outcome.cycle_id),
                    column=column,
                    field_name=field_name,
                    old_value=cell.value,
                    new_value=desired,
                    row_index=row_index,
                )
            )
            if not self._dry_run:
                was_empty = _is_empty(cell.value)
                cell.value = desired
                self._apply_number_format(column, cell, was_empty=was_empty)
        if outcome.status == "shadow_rejected":
            writes.extend(self._write_shadow_display_fields(outcome, row_index))
        return writes

    def _write_main_log_fields_for_row(
        self,
        outcome: TradeOutcome,
        row_index: int,
    ) -> list[SystemFieldWritten]:
        writes: list[SystemFieldWritten] = []
        for column, field_name in MAIN_LOG_COLUMNS.items():
            cell = self.main_sheet[f"{column}{row_index}"]
            desired = self._system_value(outcome, column, field_name)
            if self._cell_matches_system_value(
                column,
                cell.value,
                desired,
                date_columns=MAIN_LOG_DATE_COLUMNS,
            ):
                continue

            if not _is_empty(cell.value):
                self._report.system_overwrites.append(
                    SystemOverwrite(
                        underlying=outcome.underlying,
                        cycle_id_short=_short_cycle_id(outcome.cycle_id),
                        field_name=field_name,
                        excel_value=cell.value,
                        storage_value=desired,
                    )
                )

            writes.append(
                SystemFieldWritten(
                    underlying=outcome.underlying,
                    cycle_id_short=_short_cycle_id(outcome.cycle_id),
                    column=column,
                    field_name=field_name,
                    old_value=cell.value,
                    new_value=desired,
                    row_index=row_index,
                )
            )
            if not self._dry_run:
                was_empty = _is_empty(cell.value)
                cell.value = desired
                self._apply_main_log_number_format(column, cell, was_empty=was_empty)
        return writes

    def _write_shadow_display_fields(
        self,
        outcome: TradeOutcome,
        row_index: int,
    ) -> list[SystemFieldWritten]:
        writes: list[SystemFieldWritten] = []
        desired_values = {
            "O": ("entry_triggered", "NO (shadow)"),
            "W": ("thesis_played_out", None),
            "X": ("win_source", None),
            "Y": ("system_contribution", None),
            "AG": ("audit_notes", None),
        }
        for column, (field_name, desired) in desired_values.items():
            cell = self.sheet[f"{column}{row_index}"]
            if self._cell_matches_system_value(column, cell.value, desired):
                continue
            writes.append(
                SystemFieldWritten(
                    underlying=outcome.underlying,
                    cycle_id_short=_short_cycle_id(outcome.cycle_id),
                    column=column,
                    field_name=field_name,
                    old_value=cell.value,
                    new_value=desired,
                    row_index=row_index,
                )
            )
            if not self._dry_run:
                cell.value = desired
        return writes

    def _write_rejected_candidate_row(
        self,
        entry: DecisionLogEntry,
        row_index: int,
        representative: TradeOutcome,
    ) -> None:
        for column in ALL_LOG_COLUMNS:
            self.sheet.cell(row_index, column).value = None

        cycle_date = _date_value(representative.cycle_date) or entry.created_at.date()
        values = {
            "A": cycle_date,
            "B": _short_cycle_id(entry.cycle_id),
            "D": entry.candidate,
            "H": entry.conviction_rating,
        }
        for column, value in values.items():
            cell = self.sheet[f"{column}{row_index}"]
            cell.value = value
            self._apply_number_format(column, cell, was_empty=True)

    def _rejection_reason(self, entry: DecisionLogEntry) -> str:
        summary = entry.summary[:120]
        if entry.rule_applied and summary:
            return f"{entry.rule_applied}: {summary}"
        return entry.rule_applied or summary

    def _shadow_rejection_reason(self, outcome: TradeOutcome) -> str | None:
        if outcome.status != "shadow_rejected":
            return None
        notes = outcome.audit_notes or ""
        prefix = "SYSTEM REJECTED:"
        if notes.startswith(prefix):
            return notes[len(prefix):].strip()
        return notes or None

    def _decision_log_key_variants(
        self,
        entry: DecisionLogEntry,
    ) -> set[tuple[str, str]]:
        cycle_id = entry.cycle_id
        return {
            _key(_short_cycle_id(cycle_id), entry.candidate),
            _key(cycle_id[:8], entry.candidate),
            _key(cycle_id, entry.candidate),
        }

    def _is_trade_log_data_row(self, row_index: int) -> bool:
        cycle_short = _cell_text(self.sheet[f"B{row_index}"].value)
        underlying = _cell_text(self.sheet[f"D{row_index}"].value)
        if not cycle_short or not underlying:
            return False
        if cycle_short.lower() in {
            "cycle id",
            "last 8 chars of cycle uuid is fine",
        }:
            return False
        if underlying.lower() in {"ticker", "underlying"}:
            return False
        return True

    def _is_main_log_data_row(self, row_index: int) -> bool:
        cycle_short = _cell_text(self.main_sheet[f"B{row_index}"].value)
        underlying = _cell_text(self.main_sheet[f"D{row_index}"].value)
        if not cycle_short or not underlying:
            return False
        if cycle_short.lower() == "cycle id":
            return False
        if underlying.lower() in {"ticker", "underlying"}:
            return False
        return True

    def _ensure_main_log_rows_for_outcomes(self) -> None:
        if self._dry_run:
            return
        for outcome in sorted(
            self._outcomes,
            key=lambda item: (
                item.cycle_date,
                _short_cycle_id(item.cycle_id),
                item.underlying,
            ),
        ):
            key = self._outcome_key(outcome)
            if key in self._main_all_row_by_key or key in self._ambiguous_outcome_keys:
                continue
            self._append_main_log_outcome_row(outcome)

    def _append_main_log_outcome_row(self, outcome: TradeOutcome) -> int | None:
        key = self._outcome_key(outcome)
        if key in self._main_all_row_by_key:
            return self._main_all_row_by_key[key]
        row_index = self._main_first_empty_data_row()
        self._copy_main_log_row_style(self._main_template_row(), row_index)
        self._write_main_log_fields_for_row(outcome, row_index)
        self._main_all_row_by_key[key] = row_index
        self._main_row_by_key[key] = row_index
        return row_index

    def _append_main_log_rejected_candidate_row(
        self,
        entry: DecisionLogEntry,
        representative: TradeOutcome,
    ) -> int | None:
        key = _key(_short_cycle_id(entry.cycle_id), entry.candidate)
        if key in self._main_all_row_by_key:
            return self._main_all_row_by_key[key]
        row_index = self._main_first_empty_data_row()
        self._copy_main_log_row_style(self._main_template_row(), row_index)
        for column in MAIN_LOG_OWNED_COLUMN_NUMBERS:
            self.main_sheet.cell(row_index, column).value = None

        cycle_date = _date_value(representative.cycle_date) or entry.created_at.date()
        values = {
            "A": cycle_date,
            "B": _short_cycle_id(entry.cycle_id),
            "C": representative.priority_theme,
            "D": entry.candidate,
            "M": "REJECTED",
        }
        for column, value in values.items():
            cell = self.main_sheet[f"{column}{row_index}"]
            cell.value = value
            self._apply_main_log_number_format(column, cell, was_empty=True)
        self._main_all_row_by_key[key] = row_index
        return row_index

    def _system_value(
        self,
        outcome: TradeOutcome,
        column: str,
        field_name: str,
    ) -> Any:
        if field_name == "cycle_id":
            return _short_cycle_id(outcome.cycle_id)
        if field_name in {"cycle_date", "entry_date", "exit_date"}:
            return _date_value(getattr(outcome, field_name))
        if field_name == "last_price_update":
            return _datetime_date_value(outcome.last_price_update)
        if field_name == "entry_instrument_price":
            return (
                outcome.entry_instrument_price
                if outcome.entry_instrument_price is not None
                else outcome.entry_underlying_price
            )
        if field_name == "exit_instrument_price":
            return (
                outcome.exit_instrument_price
                if outcome.exit_instrument_price is not None
                else outcome.exit_underlying_price
            )
        if field_name == "realized_pnl_pct":
            return (
                outcome.realized_pnl_pct
                if outcome.status.startswith("closed_")
                else None
            )
        if field_name == "current_price_display":
            return _resolve_current_price_display(outcome)
        if field_name == "pipeline_verdict":
            return "REJECTED" if outcome.status == "shadow_rejected" else "ACCEPTED"
        if field_name == "pipeline_rejection_reason":
            return self._shadow_rejection_reason(outcome)
        return getattr(outcome, field_name)

    def _apply_status_transitions(
        self,
        *,
        outcome: TradeOutcome,
        updates: dict[str, Any],
        row_index: int,
        now: datetime,
    ) -> None:
        status = updates.get("status", outcome.status)
        decision = updates.get("user_decision", outcome.user_decision)
        if decision in {"TAKE", "SKIP", "WATCH"}:
            updates.setdefault("user_decision_at", now)
        if decision == "TAKE" and outcome.status == "proposed":
            status = "watching"
        elif decision == "SKIP" and outcome.status in {"proposed", "watching"}:
            status = "skipped"
        if status != outcome.status:
            updates["status"] = status

        entry_triggered = updates.get("entry_triggered", outcome.entry_triggered)
        if entry_triggered is True and status == "watching":
            self._add_warning(
                outcome.underlying,
                _short_cycle_id(outcome.cycle_id),
                (
                    "entry_triggered=YES but status is 'watching' - "
                    "run `enter` to record entry price"
                ),
            )

        exit_reason = updates.get("exit_reason", outcome.exit_reason)
        if not _is_empty(exit_reason) and status == "open":
            self._add_warning(
                outcome.underlying,
                _short_cycle_id(outcome.cycle_id),
                (
                    "exit_reason is set but status is 'open' - "
                    "run `close` to record exit price"
                ),
            )

    def _normalize_user_value(
        self,
        field_name: str,
        value: Any,
        *,
        outcome: TradeOutcome,
        row_index: int,
    ) -> Any:
        raw = _cell_text(value)
        if field_name == "user_decision":
            normalized = raw.upper()
            if normalized in {"TAKE", "SKIP", "WATCH"}:
                return normalized
            return self._invalid_user_value(outcome, row_index, field_name, raw)
        if field_name == "entry_triggered":
            if isinstance(value, bool):
                return value
            normalized = raw.upper()
            if normalized in {"YES", "Y", "TRUE", "1"}:
                return True
            if normalized in {"NO", "N", "FALSE", "0"}:
                return False
            return self._invalid_user_value(outcome, row_index, field_name, raw)
        if field_name == "thesis_played_out":
            normalized = raw.upper()
            if normalized in {"YES", "PARTIAL", "NO"}:
                return normalized
            return self._invalid_user_value(outcome, row_index, field_name, raw)
        if field_name == "win_source":
            normalized = raw.lower()
            if normalized in {"thesis", "direction", "timing", "luck", "sizing"}:
                return normalized
            return self._invalid_user_value(outcome, row_index, field_name, raw)
        if field_name == "system_contribution":
            normalized = raw.upper()
            if normalized in {"STRONG", "NEUTRAL", "WEAK"}:
                return normalized
            return self._invalid_user_value(outcome, row_index, field_name, raw)
        return raw

    def _invalid_user_value(
        self,
        outcome: TradeOutcome,
        row_index: int,
        field_name: str,
        raw: str,
    ) -> object:
        self._add_warning(
            outcome.underlying,
            _short_cycle_id(outcome.cycle_id),
            f"ignored invalid {field_name} value {raw!r} at row {row_index}",
        )
        return _INVALID

    def _cell_matches_system_value(
        self,
        column: str,
        current: Any,
        desired: Any,
        *,
        date_columns: set[str] = DATE_COLUMNS,
    ) -> bool:
        if desired is None:
            return _is_empty(current)
        if column in date_columns:
            if isinstance(current, datetime):
                return current.date() == desired
            if isinstance(current, date):
                return current == desired
            return False
        if isinstance(desired, (float, int)):
            try:
                return abs(float(current) - float(desired)) < 1e-9
            except (TypeError, ValueError):
                return False
        return current == desired

    def _values_equal(self, field_name: str, current: Any, desired: Any) -> bool:
        if isinstance(current, datetime) and isinstance(desired, datetime):
            return current == desired
        if field_name == "entry_triggered":
            return current is desired
        return current == desired

    def _apply_number_format(self, column: str, cell, *, was_empty: bool) -> None:
        if column in DATE_COLUMNS:
            cell.number_format = "YYYY-MM-DD"
        elif column in PERCENT_COLUMNS and cell.number_format == "General":
            cell.number_format = self._column_format(column) or "0.00%"

    def _apply_main_log_number_format(self, column: str, cell, *, was_empty: bool) -> None:
        if column in MAIN_LOG_COLUMN_FORMATS:
            if was_empty and cell.number_format in {None, "General"}:
                cell.number_format = MAIN_LOG_COLUMN_FORMATS[column]
        elif column in MAIN_LOG_DATE_COLUMNS:
            cell.number_format = "YYYY-MM-DD"
        elif column in MAIN_LOG_PERCENT_COLUMNS and cell.number_format == "General":
            cell.number_format = "0.00%"

    def _column_format(self, column: str) -> str | None:
        for row_index in range(DATA_START_ROW, self.sheet.max_row + 1):
            fmt = self.sheet[f"{column}{row_index}"].number_format
            if fmt and fmt != "General":
                return fmt
        return None

    def _first_empty_data_row(self) -> int:
        for row_index in range(DATA_START_ROW, self.sheet.max_row + 2):
            if all(
                _is_empty(self.sheet.cell(row_index, column).value)
                for column in ALL_LOG_COLUMNS
            ):
                return row_index
        return self.sheet.max_row + 1

    def _main_first_empty_data_row(self) -> int:
        for row_index in range(MAIN_LOG_DATA_START_ROW, self.main_sheet.max_row + 2):
            if all(
                _is_empty(self.main_sheet.cell(row_index, column).value)
                for column in MAIN_LOG_OWNED_COLUMN_NUMBERS
            ):
                return row_index
        return self.main_sheet.max_row + 1

    def _template_row(self) -> int:
        return max(
            DATA_START_ROW,
            min(self._first_empty_data_row() - 1, self.sheet.max_row),
        )

    def _main_template_row(self) -> int:
        return max(
            MAIN_LOG_DATA_START_ROW,
            min(self._main_first_empty_data_row() - 1, self.main_sheet.max_row),
        )

    def _copy_row_style(self, source_row: int, target_row: int) -> None:
        for column in ALL_LOG_COLUMNS:
            source = self.sheet.cell(source_row, column)
            target = self.sheet.cell(target_row, column)
            self._copy_cell_display_style(source, target)

    def _copy_main_log_row_style(self, source_row: int, target_row: int) -> None:
        for column in MAIN_LOG_OWNED_COLUMN_NUMBERS:
            source = self.main_sheet.cell(source_row, column)
            target = self.main_sheet.cell(target_row, column)
            self._copy_cell_display_style(source, target)

    def _copy_cell_display_style(self, source, target) -> None:
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)

    def _outcome_key(self, outcome: TradeOutcome) -> tuple[str, str]:
        return _key(_short_cycle_id(outcome.cycle_id), outcome.underlying)

    def _add_warning(
        self,
        underlying: str | None,
        cycle_id_short: str | None,
        message: str,
    ) -> None:
        warning = SyncWarning(
            underlying=underlying,
            cycle_id_short=cycle_id_short,
            message=message,
        )
        key = (underlying, cycle_id_short, message)
        if key in self._warning_keys:
            return
        self._warning_keys.add(key)
        if self._report is not None:
            self._report.warnings.append(warning)
        if self.log_warnings:
            logger.warning(
                "Excel sync warning (%s %s): %s",
                underlying or "-",
                cycle_id_short or "-",
                message,
            )


class _InvalidUserValue:
    pass


_INVALID = _InvalidUserValue()
