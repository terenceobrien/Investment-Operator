"""Authenticated portfolio-risk workbook endpoints."""
from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook

from api.auth import verify_clerk_token


portfolio_router = APIRouter(prefix="/api/portfolio", tags=["portfolio"])
_FACTOR_ORDER = ["MKT", "AI", "MOM", "QUAL", "VAL", "SIZE", "LOWVOL"]


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _configured_risk_source() -> Path:
    configured = os.getenv("RISK_REPORT_PATH")
    if configured:
        return Path(configured).expanduser()
    home_default = Path.home() / "risk" / "current"
    if home_default.exists() or home_default.is_symlink():
        return home_default
    return _repo_root() / "risk" / "current"


def _latest_risk_workbook() -> Path:
    source = _configured_risk_source()
    if not source.exists() and not source.is_symlink():
        raise FileNotFoundError(
            f"Risk report source not found: {source}. Set RISK_REPORT_PATH to the workbook or current pointer."
        )
    resolved = source.resolve(strict=True) if source.is_symlink() else source
    if resolved.is_file():
        return resolved
    if resolved.is_dir():
        candidates = sorted(
            [*resolved.glob("risk_report_*.xlsx"), *resolved.glob("*.xlsx")],
            key=lambda path: (path.stat().st_mtime, path.name),
            reverse=True,
        )
        if candidates:
            return candidates[0]
        raise FileNotFoundError(f"No .xlsx risk report files found in {resolved}")
    raise FileNotFoundError(f"Risk report source is not a file or directory: {resolved}")


@portfolio_router.get("/risk/latest")
def latest_portfolio_risk(user: dict = Depends(verify_clerk_token)) -> dict[str, Any]:
    """Parse the latest portfolio risk workbook into frontend-ready JSON."""
    del user
    try:
        path = _latest_risk_workbook()
        return _parse_risk_workbook(path)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not parse risk workbook: {exc}") from exc


@portfolio_router.get("/risk/latest.xlsx")
def latest_portfolio_risk_xlsx(user: dict = Depends(verify_clerk_token)) -> FileResponse:
    """Stream the source risk workbook for authenticated Excel export."""
    del user
    try:
        path = _latest_risk_workbook()
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name,
    )


def _parse_risk_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    summary_rows = _rows(wb["Summary"]) if "Summary" in wb.sheetnames else []
    breadth_rows = _rows(wb["Breadth"]) if "Breadth" in wb.sheetnames else []
    risk_rows = _rows(wb["Risk Decomposition"]) if "Risk Decomposition" in wb.sheetnames else []
    stress_rows = _rows(wb["Stress Test"]) if "Stress Test" in wb.sheetnames else []
    generated_at = _generated_at(summary_rows) or datetime.now(timezone.utc).isoformat()
    summary_kv = _kv(summary_rows)
    breadth_kv = _kv(breadth_rows)
    risk_decomp = _risk_decomposition(risk_rows)
    stress = _stress_payload(summary_kv, stress_rows)
    positions = _positions(_rows(wb["Positions"]) if "Positions" in wb.sheetnames else [])
    loadings = _factor_loadings(_rows(wb["Factor Loadings"]) if "Factor Loadings" in wb.sheetnames else [])
    return {
        "generated_at": generated_at,
        "source_workbook_name": path.name,
        "total_account_value": _num(summary_kv.get("Total account value")),
        "invested_value": _num(summary_kv.get("Invested (equities)")),
        "cash_value": _num(summary_kv.get("Cash / core")),
        "invested_fraction": _num(summary_kv.get("Invested fraction")),
        "cash_fraction": _num(summary_kv.get("Cash fraction")),
        "factor_exposures": _factor_exposures(summary_rows),
        "effective_breadth": _num(
            summary_kv.get("Effective independent bets")
            or breadth_kv.get("Effective N (eigenvalue)")
        ),
        "effective_annual_breadth": _num(
            summary_kv.get("Effective annual breadth")
            or breadth_kv.get("Effective annual breadth")
        ),
        "avg_pairwise_corr": _num(breadth_kv.get("Avg pairwise correlation")),
        "concentration_ratio": _num(breadth_kv.get("Concentration ratio")),
        "top_principal_component": _num(breadth_kv.get("Top principal component")),
        "risk_decomposition": risk_decomp,
        "stress": stress,
        "positions": positions,
        "per_name_loadings": loadings,
    }


def _rows(ws: Any) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _kv(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for row in rows:
        if row and isinstance(row[0], str) and len(row) > 1:
            out[row[0]] = row[1]
    return out


def _generated_at(rows: list[tuple[Any, ...]]) -> str | None:
    if len(rows) < 2 or not rows[1] or not isinstance(rows[1][0], str):
        return None
    text = rows[1][0]
    if text.startswith("Generated "):
        return text.removeprefix("Generated ").split("|", 1)[0].strip()
    return text


def _num(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def _find_row(rows: list[tuple[Any, ...]], label: str, *, start: int = 0) -> int | None:
    for idx in range(start, len(rows)):
        first = rows[idx][0] if rows[idx] else None
        if first == label:
            return idx
    return None


def _factor_exposures(rows: list[tuple[Any, ...]]) -> list[dict[str, float | None]]:
    start = _find_row(rows, "Factor Exposures (invested sleeve)")
    if start is None:
        return []
    out: list[dict[str, float | None]] = []
    for row in rows[start + 1 :]:
        factor = row[0] if row else None
        if factor is None:
            break
        if str(factor) not in _FACTOR_ORDER:
            continue
        out.append({"factor": str(factor), "beta": _num(row[1] if len(row) > 1 else None)})
    return out


def _positions(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    header = _find_row(rows, "Ticker")
    if header is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[header + 1 :]:
        ticker = row[0] if row else None
        if ticker is None:
            break
        text = str(ticker).upper()
        out.append(
            {
                "ticker": text,
                "weight": _num(row[1] if len(row) > 1 else None),
                "value": _num(row[2] if len(row) > 2 else None),
                "is_cash": text == "CASH",
            }
        )
    return out


def _factor_loadings(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    header = _find_row(rows, "ticker")
    if header is None:
        return []
    headers = [str(value) if value is not None else "" for value in rows[header]]
    out: list[dict[str, Any]] = []
    for row in rows[header + 1 :]:
        ticker = row[0] if row else None
        if ticker is None:
            break
        values = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        out.append(
            {
                "ticker": str(ticker).upper(),
                "loadings": {
                    factor: _num(values.get(factor))
                    for factor in _FACTOR_ORDER
                },
                "r2": _num(values.get("r2")),
            }
        )
    return out


def _risk_decomposition(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    block = _find_row(rows, "EWMA (recent-weighted)")
    if block is None:
        return {"total_vol": None, "factor_share": None, "specific_share": None, "factors": []}
    kv: dict[str, Any] = {}
    factor_header: int | None = None
    for idx in range(block + 1, len(rows)):
        first = rows[idx][0] if rows[idx] else None
        if first == "factor":
            factor_header = idx
            break
        if isinstance(first, str) and len(rows[idx]) > 1:
            kv[first] = rows[idx][1]
    factors: list[dict[str, Any]] = []
    if factor_header is not None:
        for row in rows[factor_header + 1 :]:
            factor = row[0] if row else None
            if factor is None:
                break
            factors.append(
                {
                    "factor": str(factor),
                    "exposure": _num(row[1] if len(row) > 1 else None),
                    "pct_of_total_var": _num(row[3] if len(row) > 3 else None),
                }
            )
    return {
        "total_vol": _num(kv.get("Total vol (annual)")),
        "factor_share": _num(kv.get("Factor share of variance")),
        "specific_share": _num(kv.get("Specific share of variance")),
        "factors": factors,
    }


def _stress_payload(summary_kv: dict[str, Any], rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    kv = _kv(rows)
    header = _find_row(rows, "factor")
    contributions: list[dict[str, Any]] = []
    if header is not None:
        for row in rows[header + 1 :]:
            factor = row[0] if row else None
            if factor is None:
                break
            contributions.append(
                {
                    "factor": str(factor),
                    "contribution": _num(row[3] if len(row) > 3 else None),
                }
            )
    return {
        "stressed_total_vol": _num(
            kv.get("Stressed total vol")
            or summary_kv.get("Stressed total vol (annual)")
        ),
        "sleeve_drawdown": _num(
            kv.get("Implied sleeve drawdown")
            or summary_kv.get("Implied sleeve drawdown")
        ),
        "whole_book_drawdown": _num(summary_kv.get("Implied whole-book drawdown")),
        "contributions": contributions,
    }
