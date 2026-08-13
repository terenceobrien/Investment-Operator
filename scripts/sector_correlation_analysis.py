#!/usr/bin/env python3
"""Build a sector-correlation stress research workbook.

The script reads a daily market-data CSV, filters to close observations when
available, computes rolling sector-correlation stress indicators for 9-sector
and 11-sector universes, appends forward drawdown/return research labels for
SPY/QQQ/IWM, and exports a formatted Excel workbook.
"""

from __future__ import annotations

import argparse
import bisect
import itertools
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = "~/data/correlation_signal_dataset.csv"
DEFAULT_OUTPUT = "sector_correlation_analysis.xlsx"

LONG_SECTORS = ["XLB", "XLE", "XLF", "XLI", "XLK", "XLP", "XLU", "XLV", "XLY"]
FULL_SECTORS = LONG_SECTORS + ["XLRE", "XLC"]
BENCHMARKS = ["SPY", "QQQ", "IWM"]
WINDOWS = [10, 20, 60]
VELOCITY_LAGS = [1, 3, 5, 10]
BREADTH_LAGS = [5, 10]
CORR_THRESHOLDS = [0.30, 0.50, 0.70]
FWD_HORIZONS = [5, 10, 21, 42, 63]
PRIMARY_WINDOW = 20
PRIMARY_VELOCITY_LAG = 5
DEFAULT_MIN_HISTORY = 252

UNIVERSES = [
    ("s9", LONG_SECTORS),
    ("s11", FULL_SECTORS),
]


@dataclass(frozen=True)
class CleanResult:
    data: pd.DataFrame
    date_column: str
    raw_rows: int
    rows_after_signal_filter: int
    invalid_date_rows: int
    duplicate_date_rows_before_filtering: int
    duplicate_date_excess_before_filtering: int
    duplicate_date_rows_after_filtering: int
    duplicate_date_excess_after_filtering: int
    final_unique_daily_observations: int
    signal_time_filter_applied: bool


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create sector correlation stress indicators and Excel workbook."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
        help=f"Input CSV path (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Output Excel workbook path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--min-history",
        type=int,
        default=DEFAULT_MIN_HISTORY,
        help="Minimum prior valid observations for expanding z-scores/percentiles.",
    )
    return parser.parse_args()


def resolve_input_path(raw_path: str) -> Path:
    """Respect the requested default while supporting the repo-local data copy."""
    path = Path(raw_path).expanduser()
    fallback = Path.cwd() / "data" / "correlation_signal_dataset.csv"
    if raw_path == DEFAULT_INPUT and not path.exists() and fallback.exists():
        print(f"Input path not found: {path}")
        print(f"Using repo-local fallback: {fallback}")
        return fallback
    return path


def find_date_column(columns: Iterable[str]) -> str:
    candidates = ["date", "trading_date", "asof_date", "datetime", "timestamp"]
    by_lower = {col.lower(): col for col in columns}
    for candidate in candidates:
        if candidate in by_lower:
            return by_lower[candidate]
    raise ValueError("Could not identify a date column. Expected a column such as 'date'.")


def find_signal_time_column(columns: Iterable[str]) -> str | None:
    by_lower = {col.lower(): col for col in columns}
    return by_lower.get("signal_time")


def clean_daily_data(input_path: Path) -> CleanResult:
    if not input_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {input_path}")

    raw = pd.read_csv(input_path)
    raw_rows = len(raw)
    date_column = find_date_column(raw.columns)

    working = raw.copy()
    working["_source_order"] = np.arange(len(working))
    try:
        working["_analysis_date"] = pd.to_datetime(
            working[date_column], errors="coerce", format="mixed"
        )
    except TypeError:
        working["_analysis_date"] = pd.to_datetime(working[date_column], errors="coerce")
    invalid_date_rows = int(working["_analysis_date"].isna().sum())
    working = working.dropna(subset=["_analysis_date"]).copy()

    duplicate_date_rows_before = int(
        working.duplicated("_analysis_date", keep=False).sum()
    )
    duplicate_date_excess_before = int(
        working.duplicated("_analysis_date", keep="first").sum()
    )

    signal_time_column = find_signal_time_column(working.columns)
    signal_time_filter_applied = signal_time_column is not None
    if signal_time_column:
        signal_time = working[signal_time_column].astype(str).str.strip().str.lower()
        working = working.loc[signal_time.eq("close")].copy()
        if working.empty:
            raise ValueError(
                "Column 'signal_time' exists, but no rows have signal_time == 'close'."
            )

    rows_after_signal_filter = len(working)
    duplicate_date_rows_after = int(
        working.duplicated("_analysis_date", keep=False).sum()
    )
    duplicate_date_excess_after = int(
        working.duplicated("_analysis_date", keep="first").sum()
    )

    # If multiple close rows remain for a date, keep the last source row and report it.
    working = (
        working.sort_values(["_analysis_date", "_source_order"])
        .drop_duplicates("_analysis_date", keep="last")
        .sort_values("_analysis_date")
        .reset_index(drop=True)
    )

    cleaned = working.drop(columns=[date_column], errors="ignore")
    cleaned.insert(0, "Date", working["_analysis_date"].to_numpy())
    cleaned = cleaned.drop(columns=["_analysis_date", "_source_order"], errors="ignore")

    return CleanResult(
        data=cleaned,
        date_column=date_column,
        raw_rows=raw_rows,
        rows_after_signal_filter=rows_after_signal_filter,
        invalid_date_rows=invalid_date_rows,
        duplicate_date_rows_before_filtering=duplicate_date_rows_before,
        duplicate_date_excess_before_filtering=duplicate_date_excess_before,
        duplicate_date_rows_after_filtering=duplicate_date_rows_after,
        duplicate_date_excess_after_filtering=duplicate_date_excess_after,
        final_unique_daily_observations=int(cleaned["Date"].nunique()),
        signal_time_filter_applied=signal_time_filter_applied,
    )


def column_tokens(column: str) -> list[str]:
    return [tok for tok in re.split(r"[^a-zA-Z0-9]+", column.lower()) if tok]


def ticker_matches(tokens: list[str], ticker: str) -> bool:
    ticker_key = ticker.lower().replace("^", "")
    ticker_parts = [tok for tok in re.split(r"[^a-zA-Z0-9]+", ticker_key) if tok]
    if ticker_key in tokens:
        return True
    return bool(ticker_parts) and all(part in tokens for part in ticker_parts)


def has_non_daily_horizon(tokens: list[str]) -> bool:
    horizon_tokens = [tok for tok in tokens if re.fullmatch(r"\d+d", tok)]
    return any(tok != "1d" for tok in horizon_tokens)


def identify_return_column(columns: Iterable[str], ticker: str) -> str | None:
    ticker_l = ticker.lower()
    candidates: list[tuple[float, str]] = []

    preferred_order = [
        f"ret_{ticker_l}_1d",
        f"{ticker_l}_ret_1d",
        f"{ticker_l}_return_1d",
        f"return_{ticker_l}_1d",
        f"daily_return_{ticker_l}",
        f"{ticker_l}_daily_return",
        f"{ticker_l}_ret",
        f"{ticker_l}_return",
    ]
    preferred_scores = {
        pattern: 100.0 - idx * 3.0 for idx, pattern in enumerate(preferred_order)
    }

    for column in columns:
        lower = column.lower()
        tokens = column_tokens(column)
        if not ticker_matches(tokens, ticker):
            continue
        if any(tok in {"fwd", "forward", "future", "lead", "target", "outcome"} for tok in tokens):
            continue
        if has_non_daily_horizon(tokens):
            continue

        is_return_like = (
            any(tok in {"ret", "return", "returns"} for tok in tokens)
            or "return" in lower
        )

        if is_return_like:
            score = preferred_scores.get(lower, 40.0)
            if "1d" in tokens:
                score += 20.0
            if "daily" in tokens:
                score += 10.0
            if tokens and tokens[0] == "ret":
                score += 5.0
            if tokens and tokens[0] == ticker_l:
                score += 2.0
            score -= len(tokens) * 0.01
            candidates.append((score, column))
        elif lower == ticker_l:
            # Some research files use the ticker itself as a return column.
            candidates.append((1.0, column))

    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], item[1].lower()))
    return candidates[0][1]


def identify_required_return_columns(
    data: pd.DataFrame,
    tickers: Iterable[str],
    label: str,
) -> dict[str, str]:
    mapping: dict[str, str] = {}
    missing: list[str] = []
    for ticker in tickers:
        column = identify_return_column(data.columns, ticker)
        if column is None:
            missing.append(ticker)
        else:
            mapping[ticker] = column

    if missing:
        print(f"ERROR: Missing expected {label} daily return series:")
        for ticker in missing:
            print(f"  - {ticker}")
        print("No missing return series were substituted with price columns.")
        raise SystemExit(1)

    return mapping


def identify_price_column(columns: Iterable[str], ticker: str) -> str | None:
    candidates: list[tuple[float, str]] = []
    for column in columns:
        lower = column.lower()
        tokens = column_tokens(column)
        if not ticker_matches(tokens, ticker):
            continue
        if any(tok in {"ret", "return", "returns", "fwd", "forward"} for tok in tokens):
            continue
        if any(tok in {"prev", "previous"} for tok in tokens):
            continue
        if any(tok in {"volume", "vol", "range", "clv", "z"} for tok in tokens):
            continue

        score = 0.0
        if "adj" in tokens and "close" in tokens:
            score = 100.0
        elif "adjusted" in tokens and "close" in tokens:
            score = 100.0
        elif "adjclose" in tokens:
            score = 98.0
        elif "close" in tokens:
            score = 80.0
        elif lower == ticker.lower():
            score = 10.0

        if score:
            score -= len(tokens) * 0.01
            candidates.append((score, column))

    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], item[1].lower()))
    return candidates[0][1]


def identify_vix_column(columns: Iterable[str]) -> str | None:
    lower_map = {col.lower(): col for col in columns}
    for candidate in ["vix_level", "vix_close", "vix", "^vix_close", "^vix"]:
        if candidate in lower_map:
            return lower_map[candidate]

    scored: list[tuple[int, str]] = []
    for column in columns:
        tokens = column_tokens(column)
        if "vix" not in tokens:
            continue
        if any(tok in {"ret", "return", "returns", "z", "change", "slope"} for tok in tokens):
            continue
        score = 10
        if "level" in tokens:
            score += 5
        if "close" in tokens:
            score += 4
        scored.append((score, column))

    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], item[1].lower()))
    return scored[0][1]


def build_return_frame(data: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
    return pd.DataFrame(
        {ticker: pd.to_numeric(data[column], errors="coerce") for ticker, column in mapping.items()},
        index=data.index,
    )


def pair_column(prefix: str, window: int, first: str, second: str) -> str:
    return f"{prefix}_corr{window}_{first}_{second}"


def complete_window_mask(returns: pd.DataFrame, window: int) -> pd.Series:
    complete_daily_rows = returns.notna().all(axis=1).astype(float)
    return complete_daily_rows.rolling(window, min_periods=window).sum().eq(window)


def calculate_rolling_pair_correlations(
    returns: pd.DataFrame,
    sectors: list[str],
    prefix: str,
    windows: Iterable[int],
) -> tuple[pd.DataFrame, list[tuple[str, str]]]:
    pairs = list(itertools.combinations(sectors, 2))
    columns: dict[str, pd.Series] = {}

    for window in windows:
        mask = complete_window_mask(returns[sectors], window)
        for first, second in pairs:
            col = pair_column(prefix, window, first, second)
            columns[col] = (
                returns[first]
                .rolling(window, min_periods=window)
                .corr(returns[second])
                .where(mask)
            )

    return pd.DataFrame(columns, index=returns.index), pairs


def threshold_suffix(threshold: float) -> str:
    return f"{int(round(threshold * 100)):03d}"


def calculate_aggregate_correlation_stats(
    pair_corrs: pd.DataFrame,
    pairs: list[tuple[str, str]],
    prefix: str,
    windows: Iterable[int],
) -> pd.DataFrame:
    output = pd.DataFrame(index=pair_corrs.index)
    pair_count = len(pairs)

    for window in windows:
        cols = [pair_column(prefix, window, first, second) for first, second in pairs]
        frame = pair_corrs[cols]
        full_valid = frame.notna().sum(axis=1).eq(pair_count)

        output[f"{prefix}_avg_corr_{window}d"] = frame.mean(axis=1).where(full_valid)
        output[f"{prefix}_corr{window}_pair_std"] = (
            frame.std(axis=1, ddof=0).where(full_valid)
        )
        output[f"{prefix}_corr{window}_pair_min"] = frame.min(axis=1).where(full_valid)
        output[f"{prefix}_corr{window}_pair_max"] = frame.max(axis=1).where(full_valid)

        for threshold in CORR_THRESHOLDS:
            suffix = threshold_suffix(threshold)
            output[f"{prefix}_corr{window}_pct_above_{suffix}"] = (
                frame.gt(threshold).sum(axis=1) / pair_count * 100.0
            ).where(full_valid)

    return output


def add_average_correlation_velocity(
    stats: pd.DataFrame,
    prefix: str,
    windows: Iterable[int],
    lags: Iterable[int],
) -> pd.DataFrame:
    output = stats.copy()
    for window in windows:
        avg_col = f"{prefix}_avg_corr_{window}d"
        for lag in lags:
            output[f"{prefix}_avg_corr{window}_vel_{lag}d"] = (
                output[avg_col] - output[avg_col].shift(lag)
            )

    output[f"{prefix}_avg_corr_10d_minus_60d"] = (
        output[f"{prefix}_avg_corr_10d"] - output[f"{prefix}_avg_corr_60d"]
    )
    output[f"{prefix}_avg_corr_20d_minus_60d"] = (
        output[f"{prefix}_avg_corr_20d"] - output[f"{prefix}_avg_corr_60d"]
    )
    return output


def calculate_rising_correlation_breadth(
    pair_corrs: pd.DataFrame,
    pairs: list[tuple[str, str]],
    prefix: str,
    windows: Iterable[int],
    lags: Iterable[int],
) -> pd.DataFrame:
    output = pd.DataFrame(index=pair_corrs.index)
    pair_count = len(pairs)

    for window in windows:
        cols = [pair_column(prefix, window, first, second) for first, second in pairs]
        frame = pair_corrs[cols]
        for lag in lags:
            prior = frame.shift(lag)
            valid_delta = frame.notna() & prior.notna()
            full_valid_delta = valid_delta.sum(axis=1).eq(pair_count)
            rising = frame.gt(prior) & valid_delta
            output[f"{prefix}_corr{window}_breadth_rising_{lag}d"] = (
                rising.sum(axis=1) / pair_count * 100.0
            ).where(full_valid_delta)

    return output


def calculate_pca_stats(
    returns: pd.DataFrame,
    sectors: list[str],
    prefix: str,
    windows: Iterable[int],
) -> pd.DataFrame:
    output = pd.DataFrame(index=returns.index)
    values = returns[sectors].to_numpy(dtype=float)
    sector_count = len(sectors)

    for window in windows:
        eigenvalues = np.full(len(returns), np.nan, dtype=float)
        shares = np.full(len(returns), np.nan, dtype=float)

        for end_idx in range(window - 1, len(returns)):
            block = values[end_idx - window + 1 : end_idx + 1, :]
            if block.shape != (window, sector_count) or np.isnan(block).any():
                continue

            try:
                corr_matrix = np.corrcoef(block, rowvar=False)
                if corr_matrix.shape != (sector_count, sector_count):
                    continue
                if not np.isfinite(corr_matrix).all():
                    continue
                eigvals = np.linalg.eigvalsh(corr_matrix)
                largest = float(np.max(eigvals))
            except np.linalg.LinAlgError:
                continue
            except FloatingPointError:
                continue

            eigenvalues[end_idx] = largest
            shares[end_idx] = largest / sector_count

        eigen_col = f"{prefix}_pc1_eigenvalue_{window}d"
        share_col = f"{prefix}_pc1_share_{window}d"
        output[eigen_col] = eigenvalues
        output[share_col] = shares
        for lag in [5, 10]:
            output[f"{share_col}_vel_{lag}d"] = output[share_col] - output[share_col].shift(lag)

    return output


def expanding_prior_zscore(series: pd.Series, min_history: int) -> pd.Series:
    prior = series.shift(1)
    counts = prior.expanding(min_periods=min_history).count()
    means = prior.expanding(min_periods=min_history).mean()
    stds = prior.expanding(min_periods=min_history).std(ddof=1)
    zscores = (series - means) / stds
    zscores = zscores.where(counts.ge(min_history) & stds.gt(0))
    return zscores


def expanding_prior_percentile(series: pd.Series, min_history: int) -> pd.Series:
    output = np.full(len(series), np.nan, dtype=float)
    history: list[float] = []

    for idx, value in enumerate(series.to_numpy(dtype=float)):
        if np.isnan(value):
            continue
        if len(history) >= min_history:
            rank = bisect.bisect_right(history, float(value))
            output[idx] = rank / len(history) * 100.0
        bisect.insort(history, float(value))

    return pd.Series(output, index=series.index)


def add_expanding_normalization(
    stats: pd.DataFrame,
    prefix: str,
    min_history: int,
) -> pd.DataFrame:
    output = stats.copy()
    primary_series = [
        f"{prefix}_avg_corr20_vel_5d",
        f"{prefix}_avg_corr20_vel_10d",
        f"{prefix}_pc1_share_20d_vel_5d",
        f"{prefix}_pc1_share_20d_vel_10d",
    ]

    for column in primary_series:
        if column not in output:
            continue
        output[f"{column}_z"] = expanding_prior_zscore(output[column], min_history)
        output[f"{column}_pct"] = expanding_prior_percentile(output[column], min_history)

    return output


def add_stress_score_and_events(stats: pd.DataFrame, prefix: str) -> pd.DataFrame:
    output = stats.copy()
    velocity_pct = f"{prefix}_avg_corr20_vel_5d_pct"
    breadth = f"{prefix}_corr20_breadth_rising_5d"
    pc1_pct = f"{prefix}_pc1_share_20d_vel_5d_pct"

    component_cols = [
        f"{prefix}_stress_component_avg_corr20_vel_5d_pct",
        f"{prefix}_stress_component_corr20_breadth_rising_5d",
        f"{prefix}_stress_component_pc1_share20_vel_5d_pct",
    ]
    output[component_cols[0]] = (output[velocity_pct] / 100.0).clip(0, 1)
    output[component_cols[1]] = (output[breadth] / 100.0).clip(0, 1)
    output[component_cols[2]] = (output[pc1_pct] / 100.0).clip(0, 1)
    output[f"{prefix}_systemic_corr_stress"] = output[component_cols].mean(axis=1)

    for threshold in [90, 95, 99]:
        output[f"{prefix}_avg_corr20_vel_5d_pct_gte_{threshold}_flag"] = (
            output[velocity_pct].ge(threshold).fillna(False).astype("int8")
        )

    output[f"{prefix}_vel95_breadth80_flag"] = (
        output[velocity_pct].ge(95).fillna(False)
        & output[breadth].ge(80).fillna(False)
    ).astype("int8")
    output[f"{prefix}_vel99_breadth80_flag"] = (
        output[velocity_pct].ge(99).fillna(False)
        & output[breadth].ge(80).fillna(False)
    ).astype("int8")

    return output


def calculate_forward_outcomes(
    returns: pd.Series,
    ticker: str,
    horizons: Iterable[int],
    prices: pd.Series | None = None,
) -> pd.DataFrame:
    returns = pd.to_numeric(returns, errors="coerce")
    prices = pd.to_numeric(prices, errors="coerce") if prices is not None else None
    output = pd.DataFrame(index=returns.index)
    n = len(returns)

    for horizon in horizons:
        terminal = np.full(n, np.nan, dtype=float)
        max_dd = np.full(n, np.nan, dtype=float)

        for idx in range(n):
            end_idx = idx + horizon
            if end_idx >= n:
                continue

            if prices is not None and pd.notna(prices.iloc[idx]) and prices.iloc[idx] > 0:
                future_prices = prices.iloc[idx + 1 : end_idx + 1]
                if len(future_prices) != horizon or future_prices.isna().any():
                    continue
                wealth_path = np.r_[1.0, future_prices.to_numpy(dtype=float) / float(prices.iloc[idx])]
                terminal[idx] = wealth_path[-1] - 1.0
            else:
                future_returns = returns.iloc[idx + 1 : end_idx + 1]
                if len(future_returns) != horizon or future_returns.isna().any():
                    continue
                wealth_path = np.r_[1.0, np.cumprod(1.0 + future_returns.to_numpy(dtype=float))]
                terminal[idx] = wealth_path[-1] - 1.0

            running_peak = np.maximum.accumulate(wealth_path)
            drawdowns = wealth_path / running_peak - 1.0
            max_dd[idx] = min(0.0, float(np.min(drawdowns)))

        output[f"{ticker}_fwd_{horizon}d_max_dd"] = max_dd
        output[f"{ticker}_fwd_{horizon}d_ret"] = terminal

    return output


def calculate_trailing_return(
    returns: pd.Series,
    horizon: int,
    prices: pd.Series | None = None,
) -> pd.Series:
    returns = pd.to_numeric(returns, errors="coerce")
    if prices is not None:
        prices = pd.to_numeric(prices, errors="coerce")
        return prices / prices.shift(horizon) - 1.0
    return (1.0 + returns).rolling(horizon, min_periods=horizon).apply(np.prod, raw=True) - 1.0


def first_valid_date(dates: pd.Series, series: pd.Series) -> pd.Timestamp | pd.NaT:
    valid = series.notna()
    if not valid.any():
        return pd.NaT
    return pd.to_datetime(dates.loc[valid].iloc[0])


def last_valid_date(dates: pd.Series, series: pd.Series) -> pd.Timestamp | pd.NaT:
    valid = series.notna()
    if not valid.any():
        return pd.NaT
    return pd.to_datetime(dates.loc[valid].iloc[-1])


def build_signal_columns(stats: pd.DataFrame) -> list[str]:
    ordered: list[str] = []
    for prefix, _sectors in UNIVERSES:
        ordered.extend([f"{prefix}_avg_corr_{window}d" for window in WINDOWS])
        ordered.extend(
            [
                f"{prefix}_avg_corr_10d_minus_60d",
                f"{prefix}_avg_corr_20d_minus_60d",
            ]
        )
        for window in WINDOWS:
            ordered.extend(
                [f"{prefix}_avg_corr{window}_vel_{lag}d" for lag in VELOCITY_LAGS]
            )
        ordered.extend(
            [
                f"{prefix}_avg_corr20_vel_5d_z",
                f"{prefix}_avg_corr20_vel_5d_pct",
                f"{prefix}_avg_corr20_vel_10d_z",
                f"{prefix}_avg_corr20_vel_10d_pct",
            ]
        )
        for window in WINDOWS:
            ordered.extend(
                [
                    f"{prefix}_corr{window}_breadth_rising_5d",
                    f"{prefix}_corr{window}_breadth_rising_10d",
                    f"{prefix}_corr{window}_pct_above_030",
                    f"{prefix}_corr{window}_pct_above_050",
                    f"{prefix}_corr{window}_pct_above_070",
                    f"{prefix}_corr{window}_pair_std",
                    f"{prefix}_corr{window}_pair_min",
                    f"{prefix}_corr{window}_pair_max",
                ]
            )
        for window in WINDOWS:
            ordered.extend(
                [
                    f"{prefix}_pc1_eigenvalue_{window}d",
                    f"{prefix}_pc1_share_{window}d",
                    f"{prefix}_pc1_share_{window}d_vel_5d",
                    f"{prefix}_pc1_share_{window}d_vel_10d",
                ]
            )
        ordered.extend(
            [
                f"{prefix}_pc1_share_20d_vel_5d_z",
                f"{prefix}_pc1_share_20d_vel_5d_pct",
                f"{prefix}_pc1_share_20d_vel_10d_z",
                f"{prefix}_pc1_share_20d_vel_10d_pct",
                f"{prefix}_stress_component_avg_corr20_vel_5d_pct",
                f"{prefix}_stress_component_corr20_breadth_rising_5d",
                f"{prefix}_stress_component_pc1_share20_vel_5d_pct",
                f"{prefix}_systemic_corr_stress",
                f"{prefix}_avg_corr20_vel_5d_pct_gte_90_flag",
                f"{prefix}_avg_corr20_vel_5d_pct_gte_95_flag",
                f"{prefix}_avg_corr20_vel_5d_pct_gte_99_flag",
                f"{prefix}_vel95_breadth80_flag",
                f"{prefix}_vel99_breadth80_flag",
            ]
        )

    return [col for col in ordered if col in stats.columns]


def build_pair_columns(pair_map: dict[str, list[tuple[str, str]]]) -> list[str]:
    ordered: list[str] = []
    for prefix, _sectors in UNIVERSES:
        pairs = pair_map[prefix]
        for window in WINDOWS:
            ordered.extend(
                [pair_column(prefix, window, first, second) for first, second in pairs]
            )
    return ordered


def build_forward_columns() -> list[str]:
    columns: list[str] = []
    for ticker in BENCHMARKS:
        columns.extend([f"{ticker}_fwd_{horizon}d_max_dd" for horizon in FWD_HORIZONS])
        columns.extend([f"{ticker}_fwd_{horizon}d_ret" for horizon in FWD_HORIZONS])
    return columns


def build_daily_summary(
    clean: pd.DataFrame,
    benchmark_returns: pd.DataFrame,
    stats: pd.DataFrame,
    forward_outcomes: pd.DataFrame,
    vix_column: str | None,
    spy_price: pd.Series | None,
) -> pd.DataFrame:
    summary = pd.DataFrame(index=clean.index)
    summary["Date"] = clean["Date"]
    for ticker in BENCHMARKS:
        summary[f"{ticker}_return"] = benchmark_returns[ticker]

    if vix_column:
        summary["VIX"] = pd.to_numeric(clean[vix_column], errors="coerce")

    summary["SPY_trailing_5d_ret"] = calculate_trailing_return(
        benchmark_returns["SPY"], 5, spy_price
    )
    summary["SPY_trailing_10d_ret"] = calculate_trailing_return(
        benchmark_returns["SPY"], 10, spy_price
    )

    signal_columns = build_signal_columns(stats)
    summary = pd.concat([summary, stats[signal_columns]], axis=1)
    summary = pd.concat([summary, forward_outcomes[build_forward_columns()]], axis=1)
    return summary.replace([np.inf, -np.inf], np.nan)


def build_extreme_events(daily_summary: pd.DataFrame) -> pd.DataFrame:
    s9_pct = daily_summary.get("s9_avg_corr20_vel_5d_pct")
    s11_pct = daily_summary.get("s11_avg_corr20_vel_5d_pct")
    trigger = pd.Series(False, index=daily_summary.index)
    if s9_pct is not None:
        trigger = trigger | s9_pct.ge(95).fillna(False)
    if s11_pct is not None:
        trigger = trigger | s11_pct.ge(95).fillna(False)

    columns = ["Date"]
    if "VIX" in daily_summary:
        market_context = ["SPY_return", "SPY_trailing_5d_ret", "SPY_trailing_10d_ret", "VIX"]
    else:
        market_context = ["SPY_return", "SPY_trailing_5d_ret", "SPY_trailing_10d_ret"]

    for prefix, _sectors in UNIVERSES:
        columns.extend(
            [
                f"{prefix}_avg_corr20_vel_5d_pct",
                f"{prefix}_avg_corr_20d",
                f"{prefix}_avg_corr20_vel_5d",
                f"{prefix}_corr20_breadth_rising_5d",
                f"{prefix}_corr20_pct_above_050",
                f"{prefix}_corr20_pct_above_070",
                f"{prefix}_pc1_share_20d",
                f"{prefix}_pc1_share_20d_vel_5d",
                f"{prefix}_systemic_corr_stress",
                f"{prefix}_vel95_breadth80_flag",
                f"{prefix}_vel99_breadth80_flag",
            ]
        )

    columns.extend(market_context)
    for ticker in BENCHMARKS:
        columns.extend([f"{ticker}_fwd_{horizon}d_max_dd" for horizon in FWD_HORIZONS])

    existing = [col for col in columns if col in daily_summary.columns]
    return daily_summary.loc[trigger, existing].copy()


def build_methodology_frame(min_history: int) -> pd.DataFrame:
    rows = [
        (
            "Sector Universes",
            "s9 uses XLB, XLE, XLF, XLI, XLK, XLP, XLU, XLV, XLY. "
            "s11 adds XLRE and XLC. Each universe is calculated separately and "
            "requires a complete rolling window for all sectors in that universe.",
        ),
        (
            "Daily Dataset",
            "Rows are filtered to signal_time == close when that column exists, dates "
            "are parsed to datetime, duplicate dates are reduced to one row, and data "
            "are sorted ascending. Missing returns are not forward-filled.",
        ),
        (
            "Rolling Correlations",
            "Pearson correlations are calculated for every unique sector pair over "
            "10, 20, and 60 trading-day windows. Pair ordering follows the predefined "
            "sector order, so pairs are not duplicated.",
        ),
        (
            "Average Correlation",
            "Average correlation is the arithmetic mean of unique off-diagonal pair "
            "correlations only. Diagonal 1.0 correlations and duplicate pair entries "
            "are excluded.",
        ),
        (
            "Correlation Velocity",
            "Velocity is current average correlation minus its value 1, 3, 5, or 10 "
            "trading days earlier. Short-minus-long spreads compare 10d and 20d "
            "average correlation to 60d average correlation.",
        ),
        (
            "Correlation Breadth",
            "Rising breadth is the percent of valid pairs whose rolling correlation "
            "increased versus 5 or 10 trading days earlier. Threshold breadth is the "
            "percent of pairs above 0.30, 0.50, and 0.70. Pair dispersion, minimum, "
            "and maximum correlations are diagnostic cross-sectional statistics.",
        ),
        (
            "PCA",
            "For each complete rolling sector-return window, PCA is performed through "
            "the sector correlation matrix eigenvalues. PC1 share equals the largest "
            "eigenvalue divided by the number of sectors.",
        ),
        (
            "Normalization",
            f"Expanding z-scores and percentile ranks use only prior valid history "
            f"and require at least {min_history} prior valid observations. Percentiles "
            "are reported on a 0-100 scale.",
        ),
        (
            "Forward Drawdown",
            "Forward max drawdown uses the path from date t through the next N trading "
            "days, beginning with wealth of 1.0 and then applying t+1 through t+N "
            "returns or equivalent future prices. Values are negative decimal returns.",
        ),
        (
            "Research Labels",
            "Forward SPY, QQQ, and IWM outcomes are research labels only. They are "
            "created after signal calculations and are not included in any correlation, "
            "PCA, normalization, stress-score, or event-flag calculation.",
        ),
        (
            "Systemic Stress Score",
            "The 20d diagnostic score is an equal-weight mean of available 0-1 "
            "components: 5d average-correlation velocity percentile, 5d rising-pair "
            "breadth, and 5d PC1-share velocity percentile. It is exploratory and not "
            "an optimized trading signal.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Topic", "Description"])


def build_data_quality_frame(
    clean_result: CleanResult,
    sector_returns: pd.DataFrame,
    benchmark_returns: pd.DataFrame,
    sector_return_columns: dict[str, str],
    benchmark_return_columns: dict[str, str],
    benchmark_price_columns: dict[str, str | None],
    vix_column: str | None,
    stats: pd.DataFrame,
) -> pd.DataFrame:
    dates = clean_result.data["Date"]
    rows: list[dict[str, object]] = []

    def add_row(
        section: str,
        item: str,
        detected_column: str | None = None,
        series: pd.Series | None = None,
        value: object = None,
        notes: str = "",
    ) -> None:
        row: dict[str, object] = {
            "section": section,
            "item": item,
            "detected_column": detected_column,
            "first_valid_date": pd.NaT,
            "last_valid_date": pd.NaT,
            "valid_observations": None,
            "missing_observations": None,
            "value": value,
            "notes": notes,
        }
        if series is not None:
            row["first_valid_date"] = first_valid_date(dates, series)
            row["last_valid_date"] = last_valid_date(dates, series)
            row["valid_observations"] = int(series.notna().sum())
            row["missing_observations"] = int(series.isna().sum())
        rows.append(row)

    for ticker in FULL_SECTORS:
        add_row(
            "sector_return_coverage",
            ticker,
            detected_column=sector_return_columns[ticker],
            series=sector_returns[ticker],
        )

    for ticker in BENCHMARKS:
        add_row(
            "benchmark_return_coverage",
            ticker,
            detected_column=benchmark_return_columns[ticker],
            series=benchmark_returns[ticker],
        )
        source_note = (
            "price-based forward path"
            if benchmark_price_columns.get(ticker)
            else "return-built forward path"
        )
        add_row(
            "benchmark_forward_source",
            ticker,
            detected_column=benchmark_price_columns.get(ticker),
            value=source_note,
        )

    add_row("optional_market_context", "VIX", detected_column=vix_column)
    add_row("dataset", "raw_rows", value=clean_result.raw_rows)
    add_row("dataset", "invalid_date_rows", value=clean_result.invalid_date_rows)
    add_row(
        "dataset",
        "signal_time_filter_applied",
        value=clean_result.signal_time_filter_applied,
    )
    add_row(
        "dataset",
        "rows_after_signal_time_filter",
        value=clean_result.rows_after_signal_filter,
    )
    add_row(
        "dataset",
        "duplicated_date_rows_before_filtering",
        value=clean_result.duplicate_date_rows_before_filtering,
    )
    add_row(
        "dataset",
        "duplicated_date_excess_before_filtering",
        value=clean_result.duplicate_date_excess_before_filtering,
    )
    add_row(
        "dataset",
        "duplicated_date_rows_after_filtering",
        value=clean_result.duplicate_date_rows_after_filtering,
    )
    add_row(
        "dataset",
        "duplicated_date_excess_after_filtering",
        value=clean_result.duplicate_date_excess_after_filtering,
    )
    add_row(
        "dataset",
        "final_unique_daily_observations",
        value=clean_result.final_unique_daily_observations,
    )

    for prefix, _sectors in UNIVERSES:
        for window in WINDOWS:
            column = f"{prefix}_avg_corr_{window}d"
            add_row(
                "model_validity",
                f"{prefix}_{window}d_average_correlation",
                detected_column=column,
                series=stats[column],
            )

    return pd.DataFrame(rows)


def bounded_issues(
    frame: pd.DataFrame,
    columns: Iterable[str],
    low: float,
    high: float,
    label: str,
    eps: float = 1e-9,
) -> list[str]:
    issues: list[str] = []
    for column in columns:
        series = pd.to_numeric(frame[column], errors="coerce").dropna()
        if series.empty:
            continue
        if series.min() < low - eps or series.max() > high + eps:
            issues.append(
                f"{label} out of bounds in {column}: min={series.min():.6f}, max={series.max():.6f}"
            )
    return issues


def validation_summary(
    clean_result: CleanResult,
    pair_map: dict[str, list[tuple[str, str]]],
    stats: pd.DataFrame,
    daily_summary: pd.DataFrame,
    sector_return_columns: dict[str, str],
    benchmark_return_columns: dict[str, str],
    vix_column: str | None,
) -> dict[str, object]:
    issues: list[str] = []
    clean = clean_result.data
    one_obs = clean["Date"].nunique() == len(clean)
    if not one_obs:
        issues.append("Cleaned dataset does not have exactly one observation per date.")

    pair_counts_ok = len(pair_map["s9"]) == 36 and len(pair_map["s11"]) == 55
    if not pair_counts_ok:
        issues.append(
            f"Unexpected pair counts: s9={len(pair_map['s9'])}, s11={len(pair_map['s11'])}."
        )

    avg_cols = [
        col
        for col in stats.columns
        if re.fullmatch(r"s(9|11)_avg_corr_(10|20|60)d", col)
    ]
    breadth_cols = [
        col
        for col in stats.columns
        if "_breadth_" in col or "_pct_above_" in col
    ]
    pc1_share_cols = [
        col
        for col in stats.columns
        if re.fullmatch(r"s(9|11)_pc1_share_(10|20|60)d", col)
    ]
    fwd_dd_cols = [col for col in daily_summary.columns if col.endswith("_max_dd")]

    issues.extend(bounded_issues(stats, avg_cols, -1.0, 1.0, "Average correlation"))
    issues.extend(bounded_issues(stats, breadth_cols, 0.0, 100.0, "Breadth metric"))
    issues.extend(bounded_issues(stats, pc1_share_cols, 0.0, 1.0, "PC1 share"))
    for column in fwd_dd_cols:
        series = pd.to_numeric(daily_summary[column], errors="coerce").dropna()
        if not series.empty and series.max() > 1e-9:
            issues.append(
                f"Forward drawdown is positive in {column}: max={series.max():.6f}"
            )

    s9_first_20d = first_valid_date(daily_summary["Date"], stats["s9_avg_corr_20d"])
    s11_first_20d = first_valid_date(daily_summary["Date"], stats["s11_avg_corr_20d"])
    s9_event95 = int(stats["s9_avg_corr20_vel_5d_pct_gte_95_flag"].sum())
    s11_event95 = int(stats["s11_avg_corr20_vel_5d_pct_gte_95_flag"].sum())
    s9_event99 = int(stats["s9_avg_corr20_vel_5d_pct_gte_99_flag"].sum())
    s11_event99 = int(stats["s11_avg_corr20_vel_5d_pct_gte_99_flag"].sum())
    union95 = int(
        (
            stats["s9_avg_corr20_vel_5d_pct_gte_95_flag"].astype(bool)
            | stats["s11_avg_corr20_vel_5d_pct_gte_95_flag"].astype(bool)
        ).sum()
    )
    union99 = int(
        (
            stats["s9_avg_corr20_vel_5d_pct_gte_99_flag"].astype(bool)
            | stats["s11_avg_corr20_vel_5d_pct_gte_99_flag"].astype(bool)
        ).sum()
    )

    print("\nValidation Summary")
    print(f"1. One observation per date: {'OK' if one_obs else 'ISSUE'}")
    print(f"   Cleaned rows={len(clean)}, unique dates={clean['Date'].nunique()}")
    print(
        f"2. Pair counts: {'OK' if pair_counts_ok else 'ISSUE'} "
        f"(s9={len(pair_map['s9'])}, s11={len(pair_map['s11'])})"
    )
    print(f"3. Average correlations bounded [-1, +1]: {'OK' if not any('Average correlation' in i for i in issues) else 'ISSUE'}")
    print(f"4. Breadth metrics bounded [0, 100]: {'OK' if not any('Breadth metric' in i for i in issues) else 'ISSUE'}")
    print(f"5. PCA first-component share bounded [0, 1]: {'OK' if not any('PC1 share' in i for i in issues) else 'ISSUE'}")
    print(f"6. Forward drawdowns <= 0: {'OK' if not any('Forward drawdown' in i for i in issues) else 'ISSUE'}")
    print(
        "7. Forward-looking data excluded from signals: OK "
        "(forward labels are appended after signal calculations)."
    )
    print(f"8. First valid s9 20d correlation model date: {format_date(s9_first_20d)}")
    print(f"   First valid s11 20d correlation model date: {format_date(s11_first_20d)}")
    print(
        "9. Stress events: "
        f"s9 >=95={s9_event95}, s11 >=95={s11_event95}, union >=95={union95}; "
        f"s9 >=99={s9_event99}, s11 >=99={s11_event99}, union >=99={union99}"
    )
    print("10. Missing expected columns/data-quality issues:")
    print("    Sector return columns:")
    for ticker in FULL_SECTORS:
        print(f"      {ticker}: {sector_return_columns[ticker]}")
    print("    Benchmark return columns:")
    for ticker in BENCHMARKS:
        print(f"      {ticker}: {benchmark_return_columns[ticker]}")
    print(f"    VIX column: {vix_column or 'not available'}")
    print(
        "    Duplicated date rows before filtering: "
        f"{clean_result.duplicate_date_rows_before_filtering} "
        f"(excess rows={clean_result.duplicate_date_excess_before_filtering})"
    )
    print(
        "    Duplicated date rows after close filtering: "
        f"{clean_result.duplicate_date_rows_after_filtering} "
        f"(excess rows={clean_result.duplicate_date_excess_after_filtering})"
    )
    if issues:
        for issue in issues:
            print(f"    ISSUE: {issue}")
    else:
        print("    No missing expected sector/index return columns and no validation bound issues.")

    return {
        "issues": issues,
        "first_valid_s9_20d": s9_first_20d,
        "first_valid_s11_20d": s11_first_20d,
        "event95_union": union95,
        "event99_union": union99,
        "event95_s9": s9_event95,
        "event95_s11": s11_event95,
        "event99_s9": s9_event99,
        "event99_s11": s11_event99,
    }


def format_date(value: object) -> str:
    if pd.isna(value):
        return "N/A"
    return pd.to_datetime(value).strftime("%Y-%m-%d")


def header_number_format(header: str) -> str | None:
    if header == "Date" or header.endswith("_date"):
        return "yyyy-mm-dd"
    if header in {"SPY_return", "QQQ_return", "IWM_return"}:
        return "0.00%"
    if header.startswith(("SPY_fwd_", "QQQ_fwd_", "IWM_fwd_")):
        return "0.00%"
    if header.startswith("SPY_trailing_"):
        return "0.00%"
    if header.endswith("_flag"):
        return "0"
    if header.endswith("_z"):
        return "0.00"
    if "_pct_above_" in header or "_breadth_" in header or header.endswith("_pct"):
        return "0.0"
    if "_corr" in header or "_pc1_" in header or "systemic_corr_stress" in header:
        return "0.0000"
    if header.startswith("VIX"):
        return "0.00"
    return None


def apply_table_format(ws, dataframe: pd.DataFrame, sheet_name: str) -> None:
    if dataframe.empty:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    for col_idx, header_cell in enumerate(ws[1], start=1):
        header = str(header_cell.value)
        fmt = header_number_format(header)
        letter = get_column_letter(col_idx)
        width = max(12, min(28, len(header) + 2))
        if sheet_name == "Pair_Correlations":
            width = max(14, min(22, len(header) + 1))
        elif sheet_name == "Methodology":
            width = 22 if col_idx == 1 else 110
        elif sheet_name == "Data_Quality":
            width = max(14, min(34, len(header) + 4))
        ws.column_dimensions[letter].width = width

        if fmt:
            for row_idx in range(2, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = fmt

    if sheet_name == "Methodology":
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 48

    if sheet_name in {"Daily_Summary", "Extreme_Events"} and ws.max_row > 2:
        for col_idx, header_cell in enumerate(ws[1], start=1):
            header = str(header_cell.value)
            if (
                header.endswith("_pct")
                or "systemic_corr_stress" in header
                or header.endswith("_flag")
            ):
                letter = get_column_letter(col_idx)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="FFFFFF",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFF2CC",
                        end_type="max",
                        end_color="F4CCCC",
                    ),
                )


def export_excel(
    output_path: Path,
    daily_summary: pd.DataFrame,
    pair_correlations: pd.DataFrame,
    extreme_events: pd.DataFrame,
    methodology: pd.DataFrame,
    data_quality: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        daily_summary.to_excel(writer, sheet_name="Daily_Summary", index=False)
        pair_correlations.to_excel(writer, sheet_name="Pair_Correlations", index=False)
        extreme_events.to_excel(writer, sheet_name="Extreme_Events", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)
        data_quality.to_excel(writer, sheet_name="Data_Quality", index=False)

        for sheet_name, dataframe in {
            "Daily_Summary": daily_summary,
            "Pair_Correlations": pair_correlations,
            "Extreme_Events": extreme_events,
            "Methodology": methodology,
            "Data_Quality": data_quality,
        }.items():
            apply_table_format(writer.book[sheet_name], dataframe, sheet_name)


def run_analysis(input_path: Path, output_path: Path, min_history: int) -> dict[str, object]:
    clean_result = clean_daily_data(input_path)
    clean = clean_result.data

    sector_return_columns = identify_required_return_columns(
        clean, FULL_SECTORS, "sector"
    )
    benchmark_return_columns = identify_required_return_columns(
        clean, BENCHMARKS, "benchmark"
    )
    sector_returns = build_return_frame(clean, sector_return_columns)
    benchmark_returns = build_return_frame(clean, benchmark_return_columns)

    benchmark_price_columns = {
        ticker: identify_price_column(clean.columns, ticker) for ticker in BENCHMARKS
    }
    vix_column = identify_vix_column(clean.columns)

    stats_parts: list[pd.DataFrame] = []
    pair_parts: list[pd.DataFrame] = []
    pair_map: dict[str, list[tuple[str, str]]] = {}

    for prefix, sectors in UNIVERSES:
        pair_corrs, pairs = calculate_rolling_pair_correlations(
            sector_returns[sectors], sectors, prefix, WINDOWS
        )
        pair_map[prefix] = pairs
        aggregate = calculate_aggregate_correlation_stats(
            pair_corrs, pairs, prefix, WINDOWS
        )
        breadth = calculate_rising_correlation_breadth(
            pair_corrs, pairs, prefix, WINDOWS, BREADTH_LAGS
        )
        pca = calculate_pca_stats(sector_returns[sectors], sectors, prefix, WINDOWS)
        stats = pd.concat([aggregate, breadth, pca], axis=1)
        stats = add_average_correlation_velocity(stats, prefix, WINDOWS, VELOCITY_LAGS)
        stats = add_expanding_normalization(stats, prefix, min_history)
        stats = add_stress_score_and_events(stats, prefix)
        stats_parts.append(stats)
        pair_parts.append(pair_corrs)

    stats_all = pd.concat(stats_parts, axis=1).replace([np.inf, -np.inf], np.nan)
    pair_correlations = pd.concat(pair_parts, axis=1)
    pair_correlations = pd.concat(
        [clean[["Date"]], pair_correlations[build_pair_columns(pair_map)]],
        axis=1,
    ).replace([np.inf, -np.inf], np.nan)

    forward_parts: list[pd.DataFrame] = []
    for ticker in BENCHMARKS:
        price_col = benchmark_price_columns.get(ticker)
        prices = pd.to_numeric(clean[price_col], errors="coerce") if price_col else None
        forward_parts.append(
            calculate_forward_outcomes(
                benchmark_returns[ticker],
                ticker,
                FWD_HORIZONS,
                prices=prices,
            )
        )
    forward_outcomes = pd.concat(forward_parts, axis=1)

    spy_price_col = benchmark_price_columns.get("SPY")
    spy_price = pd.to_numeric(clean[spy_price_col], errors="coerce") if spy_price_col else None
    daily_summary = build_daily_summary(
        clean,
        benchmark_returns,
        stats_all,
        forward_outcomes,
        vix_column,
        spy_price,
    )
    extreme_events = build_extreme_events(daily_summary)
    methodology = build_methodology_frame(min_history)
    data_quality = build_data_quality_frame(
        clean_result,
        sector_returns,
        benchmark_returns,
        sector_return_columns,
        benchmark_return_columns,
        benchmark_price_columns,
        vix_column,
        stats_all,
    )

    validation = validation_summary(
        clean_result,
        pair_map,
        stats_all,
        daily_summary,
        sector_return_columns,
        benchmark_return_columns,
        vix_column,
    )

    export_excel(
        output_path,
        daily_summary,
        pair_correlations,
        extreme_events,
        methodology,
        data_quality,
    )

    validation.update(
        {
            "dates_analyzed": len(daily_summary),
            "first_date": daily_summary["Date"].min(),
            "last_date": daily_summary["Date"].max(),
            "output_path": output_path,
            "extreme_event_rows": len(extreme_events),
        }
    )
    return validation


def main() -> None:
    args = parse_args()
    input_path = resolve_input_path(args.input)
    output_path = Path(args.output).expanduser()
    if not output_path.is_absolute():
        output_path = Path.cwd() / output_path

    result = run_analysis(input_path, output_path, args.min_history)

    print("\nFinished")
    print(f"Output path: {result['output_path']}")
    print(f"Number of dates analyzed: {result['dates_analyzed']}")
    print(
        f"Date range: {format_date(result['first_date'])} "
        f"to {format_date(result['last_date'])}"
    )
    print(f"First valid 9-sector signal date: {format_date(result['first_valid_s9_20d'])}")
    print(f"First valid 11-sector signal date: {format_date(result['first_valid_s11_20d'])}")
    print(
        "95th-percentile stress events: "
        f"{result['event95_union']} union "
        f"(s9={result['event95_s9']}, s11={result['event95_s11']})"
    )
    print(
        "99th-percentile stress events: "
        f"{result['event99_union']} union "
        f"(s9={result['event99_s9']}, s11={result['event99_s11']})"
    )
    print(f"Extreme_Events rows: {result['extreme_event_rows']}")


if __name__ == "__main__":
    main()
